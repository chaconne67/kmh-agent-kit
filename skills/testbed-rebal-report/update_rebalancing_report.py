"""서류21 리밸런싱 발생내역 엑셀 업데이트
사용: python update_rebalancing_report.py --recipe kr --schedule-id 28
(사전에 fetch_rebalancing_data.py로 JSON 데이터 생성 필요)

원칙:
- 컬럼 번호 하드코딩 없음 → 헤더에서 동적 매핑
- 수식/값 자동 분류 → 템플릿 행에서 감지
- 합계행 수식 → 기존 합계행에서 범위만 조정
- 중복 데이터 → 기존 블록 삭제 후 재작성
"""
import argparse
import json
import os
import re
import sys
from copy import copy
from datetime import datetime, date
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
GROUPS = ['안정', '중립', '적극']


# ===== 환경/설정 =====

def load_env():
    env_path = Path.home() / '.env'
    if env_path.exists():
        for line in env_path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, _, val = line.partition('=')
                os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


def get_base_dir():
    return Path(os.environ.get(
        'TESTBED_DOC_DIR',
        r'C:\Users\chaconne\Google Drive 스트리밍\내 드라이브\MOA\테스트베드2차\준비서류',
    ))


def load_recipe_config():
    config_path = SCRIPT_DIR / 'rebalancing_report_config.json'
    with open(config_path, encoding='utf-8') as f:
        config = json.load(f)
    for key, recipe in config.items():
        recipe.setdefault('dir', recipe['name'])
        recipe.setdefault('file_prefix', f'({recipe["name"]})')
    return config


def get_json_path(recipe_dir, schedule_id):
    return recipe_dir / f'rebalancing_data_{schedule_id}.json'


def sheet_names(group):
    return {
        'port': f'포트변경내역({group})',
        'mp': f'MP내역({group})',
        'balance': [f'잔고변경현황({group}{i})' for i in range(1, 4)],
    }


# ===== 엑셀 유틸리티 =====

def build_header_map(ws, header_row):
    """헤더 행에서 {헤더명: 컬럼번호} 매핑. 줄바꿈은 공백으로 치환."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is not None:
            key = str(val).replace('\n', ' ').strip()
            mapping[key] = c
    return mapping


def find_col(header_map, *candidates):
    """여러 후보 헤더명 중 매치되는 컬럼 번호. 완전일치 → 부분일치 순서."""
    for name in candidates:
        if name in header_map:
            return header_map[name]
    for name in candidates:
        for header, col in header_map.items():
            if name in header:
                return col
    raise KeyError(f'헤더를 찾을 수 없음: {candidates}\n가용: {list(header_map.keys())}')


def is_formula(cell):
    return (cell.value is not None
            and isinstance(cell.value, str)
            and cell.value.startswith('='))


def clone_formula(formula, template_row, new_row):
    """수식의 상대 행 참조를 치환. $절대 행 참조는 유지."""
    if not formula or not formula.startswith('='):
        return formula

    def replace_match(m):
        prefix = m.group(1)
        dollar = m.group(2)
        row_num = int(m.group(3))
        if dollar == '$':
            return m.group(0)
        if row_num == template_row:
            return f'{prefix}{new_row}'
        return m.group(0)

    return re.sub(r"(['\w\(\)!]*\$?[A-Z]+)(\$?)(\d+)", replace_match, formula)


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def clone_row(ws, src_row, dst_row, value_overrides=None):
    """src_row → dst_row 복제.
    - 수식 셀: 행번호 치환 후 복제
    - value_overrides 지정 셀: 해당 값으로 기입
    - 그 외 (이전 데이터): 복제하지 않음 (서식만 복제)
    """
    if value_overrides is None:
        value_overrides = {}
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if c in value_overrides:
            dst.value = value_overrides[c]
        elif is_formula(src):
            dst.value = clone_formula(src.value, src_row, dst_row)
        copy_cell_style(src, dst)


def clone_sum_row(ws, src_sum_row, dst_sum_row,
                  old_block_start, old_block_end,
                  new_block_start, new_block_end,
                  value_overrides=None):
    """합계행 복제: SUM 범위를 새 블록에 맞게 조정."""
    if value_overrides is None:
        value_overrides = {}
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_sum_row, c)
        dst = ws.cell(dst_sum_row, c)
        if c in value_overrides:
            dst.value = value_overrides[c]
        elif is_formula(src):
            formula = src.value
            # SUM/SUMIF 등의 범위 참조 조정
            # end가 old_block_end와 일치하면 같은 블록의 범위로 판단,
            # start도 무조건 new_block_start로 치환 (컬럼별 start 불일치 대응)
            def replace_range(m):
                c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                if r2 == old_block_end:
                    return f'{c1}{new_block_start}:{c2}{new_block_end}'
                nr1 = new_block_start if r1 == old_block_start else r1
                nr2 = new_block_end if r2 == old_block_end else r2
                return f'{c1}{nr1}:{c2}{nr2}'
            formula = re.sub(r'(\$?[A-Z]+)(\d+):(\$?[A-Z]+)(\d+)', replace_range, formula)
            # 합계행 자체 행참조 치환
            formula = clone_formula(formula, src_sum_row, dst_sum_row)
            dst.value = formula
        copy_cell_style(src, dst)


def to_date(val):
    """datetime/date/str → date 변환"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def find_date_col(ws, data_row):
    """데이터 행에서 datetime 값이 있는 첫 번째 컬럼 탐색"""
    for c in range(1, min(6, ws.max_column + 1)):
        val = ws.cell(data_row, c).value
        if isinstance(val, (datetime, date)):
            return c
    return None


def find_last_data_row(ws, col_idx):
    """지정 컬럼에서 값이 있는 마지막 행"""
    for row in range(ws.max_row, 0, -1):
        val = ws.cell(row, col_idx).value
        if val is not None and val != '':
            return row
    return None


def find_sum_row(ws, identifier_col, identifier_value='합계'):
    """마지막 합계행 찾기. identifier_col에 값 우선, 없으면 =SUM 수식 패턴으로 탐지."""
    # 방법 1: identifier_col에 '합계' 텍스트
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row, identifier_col).value == identifier_value:
            return row
    # 방법 2: =SUM 수식이 있는 마지막 행
    for row in range(ws.max_row, 0, -1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row, c).value
            if isinstance(val, str) and val.startswith('=SUM('):
                return row
    return None


def find_block_before_sum(ws, sum_row, date_col):
    """합계행 바로 앞의 데이터 블록 범위 (start_row, end_row)"""
    if sum_row is None or sum_row <= 1:
        return None, None
    end_row = sum_row - 1
    target = to_date(ws.cell(end_row, date_col).value)
    if target is None:
        return end_row, end_row
    start_row = end_row
    for r in range(end_row - 1, 0, -1):
        if to_date(ws.cell(r, date_col).value) == target:
            start_row = r
        else:
            break
    return start_row, end_row


def delete_block_with_sum(ws, date_col, target_date, first_data_row, sum_id_col):
    """날짜가 일치하는 블록 + 바로 뒤의 합계행까지 삭제."""
    target = to_date(target_date)
    rows_to_delete = set()
    for r in range(first_data_row, ws.max_row + 1):
        if to_date(ws.cell(r, date_col).value) == target:
            rows_to_delete.add(r)
    if rows_to_delete:
        max_row = max(rows_to_delete)
        next_row = max_row + 1
        if next_row <= ws.max_row:
            # '합계' 텍스트 또는 =SUM 수식으로 합계행 판별
            is_sum = ws.cell(next_row, sum_id_col).value == '합계'
            if not is_sum:
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(next_row, c).value
                    if isinstance(val, str) and val.startswith('=SUM('):
                        is_sum = True
                        break
            if is_sum:
                rows_to_delete.add(next_row)
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r)
    return len(rows_to_delete)


def delete_rows_by_date(ws, date_col, target_date, first_data_row):
    """날짜가 일치하는 모든 행 삭제."""
    target = to_date(target_date)
    rows = [r for r in range(first_data_row, ws.max_row + 1)
            if to_date(ws.cell(r, date_col).value) == target]
    for r in reversed(rows):
        ws.delete_rows(r)
    return len(rows)


def extract_sum_range(ws, sum_row):
    """합계행의 =SUM 수식에서 범위(start_row, end_row)를 직접 파싱."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(sum_row, c).value
        if isinstance(val, str):
            m = re.search(r'=SUM\(\$?[A-Z]+(\d+):\$?[A-Z]+(\d+)\)', val)
            if m:
                return int(m.group(1)), int(m.group(2))
    return None, None


def get_template_info(ws, date_col, isin_col):
    """템플릿으로 사용할 데이터 행과 합계행 정보 반환.
    Returns: (template_data_row, template_sum_row, old_block_start, old_block_end)
    old_block_start/end는 SUM 수식에서 직접 파싱 (날짜 추정이 아님).
    """
    sum_row = find_sum_row(ws, isin_col, '합계')
    if not sum_row:
        return None, None, None, None
    block_start, block_end = extract_sum_range(ws, sum_row)
    if block_start is None:
        # SUM 수식이 없으면 날짜 기반 폴백
        block_start, block_end = find_block_before_sum(ws, sum_row, date_col)
    template_data = block_end if block_end else sum_row - 1
    return template_data, sum_row, block_start, block_end


# ===== 시트별 편집 =====

def update_universe(ws, universe_data, existing_isins):
    """투자유니버스: 누락 종목 추가"""
    header_map = build_header_map(ws, 1)
    col = {
        'isin': find_col(header_map, '종목코드(ISIN코드)', 'ISIN코드', '종목코드'),
        'name': find_col(header_map, '종목명'),
        'market': find_col(header_map, '시장구분'),
        'asset_class': find_col(header_map, '자산군'),
        'asset_type': find_col(header_map, '자산종류'),
        'risk_grade': find_col(header_map, '위험등급'),
        'dscore': find_col(header_map, '위험도 점수', '위험도점수'),
        'danger': find_col(header_map, '위험자산여부'),
        'ticker': find_col(header_map, '비고'),
    }
    added = []
    next_row = ws.max_row + 1
    for isin, info in universe_data.items():
        if isin == 'CASH' or isin in existing_isins:
            continue
        ws.cell(next_row, col['isin'], isin)
        ws.cell(next_row, col['name'], info.get('name', ''))
        ws.cell(next_row, col['market'], info.get('market', ''))
        ws.cell(next_row, col['asset_class'], info.get('asset_class', ''))
        ws.cell(next_row, col['asset_type'], info.get('asset_type', ''))
        ws.cell(next_row, col['risk_grade'], '')
        ws.cell(next_row, col['dscore'], info.get('dscore', 0))
        ws.cell(next_row, col['danger'], 'Y' if info.get('is_danger') else 'N')
        ws.cell(next_row, col['ticker'], info.get('ticker', ''))
        added.append(isin)
        next_row += 1
    return added


def update_trade_history(ws, trades):
    """전체매매내역: 기존 동일날짜 거래 삭제 후 재작성"""
    header_map = build_header_map(ws, 1)
    date_col = find_col(header_map, '매매일자')

    # 헤더명 → JSON 키 매핑 (JSON 키도 한글)
    known = {
        '매매일자': ('매매일자', lambda v: datetime.strptime(v, '%Y-%m-%d') if isinstance(v, str) else v),
        '포트폴리오유형': ('포트폴리오유형', None),
        '일임계좌번호': ('일임계좌번호', None),
        '매매구분': ('매매구분', None),
        '종목명': ('종목명', None),
        '매매수량': ('매매수량', None),
        '매매가격': ('매매가격', None),
        '잔고수량': ('잔고수량', None),
    }
    # 종목코드 (헤더명 여러 변형 가능)
    for name in ['종목코드', 'ISIN코드', '종목코드(ISIN코드)']:
        if name in header_map:
            known[name] = ('종목코드', None)
            break

    # 매핑된 컬럼 집합
    mapped_cols = {header_map[h] for h in known if h in header_map}

    # 매핑 안 된 컬럼 = 레시피명 (헤더가 레시피마다 다름)
    col_map = {}
    for header_name, (json_key, xform) in known.items():
        if header_name in header_map:
            col_map[header_map[header_name]] = (json_key, xform)
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value is not None and c not in mapped_cols:
            col_map[c] = ('모아에셋포트명', None)
            break

    # 기존 거래 삭제 (incoming 날짜 기준)
    if trades:
        trade_dates = {to_date(datetime.strptime(t['매매일자'], '%Y-%m-%d')) for t in trades}
        rows = [r for r in range(2, ws.max_row + 1)
                if to_date(ws.cell(r, date_col).value) in trade_dates]
        for r in reversed(rows):
            ws.delete_rows(r)
        if rows:
            print(f'  기존 {len(rows)}행 삭제')

    # 새 거래 추가
    next_row = ws.max_row + 1
    for t in trades:
        for c, (json_key, xform) in col_map.items():
            val = t.get(json_key)
            if val is not None and xform:
                val = xform(val)
            ws.cell(next_row, c, val)
        next_row += 1
    return len(trades)


def update_mp(ws, mp_items, rday_str):
    """MP내역: 기존 동일 rday 블록 삭제 후 재작성"""
    rday = datetime.strptime(rday_str, '%Y-%m-%d')
    header_map = build_header_map(ws, 1)
    date_col = find_col(header_map, 'MP생성일자')
    isin_col = find_col(header_map, 'ISIN코드', 'ISIN 코드')
    ratio_col = find_col(header_map, '비중')

    # 템플릿 정보 수집 (삭제 전)
    tpl_data, tpl_sum, old_start, old_end = get_template_info(ws, date_col, isin_col)
    if not tpl_data:
        print('  [ERROR] MP 시트에 템플릿 없음')
        return 0

    # 기존 rday 블록 삭제
    deleted = delete_block_with_sum(ws, date_col, rday, 2, isin_col)
    if deleted:
        print(f'  기존 {deleted}행 삭제')
        # 삭제 후 템플릿 재탐색
        tpl_data, tpl_sum, old_start, old_end = get_template_info(ws, date_col, isin_col)
        if not tpl_data:
            print('  [ERROR] 삭제 후 템플릿 없음')
            return 0

    # 종목 정렬: CASH 마지막
    items = [i for i in mp_items if i['isincode'] != 'CASH']
    cash = [i for i in mp_items if i['isincode'] == 'CASH']
    all_items = items + cash

    # 비중 보정
    total = sum(i['ratio'] for i in all_items)
    if abs(total - 1.0) > 0.005:
        print(f'  [ERROR] 비중 합계 {total:.4f} — 허용범위(±0.5%) 초과')
        return 0
    if abs(total - 1.0) > 0.0001:
        diff = round(1.0 - total, 6)
        target = cash[0] if cash else all_items[-1]
        target['ratio'] = round(target['ratio'] + diff, 6)
        print(f'  [보정] {total:.6f} → {target["isincode"]}에 {diff:+.6f}')

    # 새 블록 시작: 템플릿 합계행 바로 다음
    start_row = tpl_sum + 1

    for i, item in enumerate(all_items):
        row = start_row + i
        clone_row(ws, tpl_data, row, {
            date_col: rday,
            isin_col: item['isincode'],
            ratio_col: item['ratio'],
        })

    # 합계행
    new_end = start_row + len(all_items) - 1
    new_sum = new_end + 1
    clone_sum_row(ws, tpl_sum, new_sum,
                  old_start, old_end, start_row, new_end,
                  {date_col: rday, isin_col: '합계'})

    return len(all_items)


def update_port_change(ws, rday_str, reason_text='정기리밸런싱'):
    """포트변경내역: 기존 동일 rday 행 삭제 후 재작성.
    date_col과 reason_col을 헤더에서 동적 탐지."""
    rday = datetime.strptime(rday_str, '%Y-%m-%d')

    # R4 헤더에서 리밸런싱 사유 컬럼 탐지
    header_r4 = build_header_map(ws, 4)
    reason_col = find_col(header_r4, '리밸런싱 사유', '리밸런싱사유', '사유')

    # date_col: 첫 번째 데이터 행에서 datetime 값이 있는 컬럼
    first_data_row = 6
    date_col = find_date_col(ws, first_data_row)
    if date_col is None:
        print('  [ERROR] 포트변경내역에서 날짜 컬럼을 찾을 수 없음')
        return False

    # 기존 rday 행 삭제
    deleted = delete_rows_by_date(ws, date_col, rday, first_data_row)
    if deleted:
        print(f'  기존 {deleted}행 삭제')

    # 템플릿 (마지막 데이터 행)
    template_row = find_last_data_row(ws, date_col)
    if not template_row or template_row < first_data_row:
        print('  [ERROR] 포트변경내역에 기존 데이터 없음')
        return False

    # 새 행 추가
    new_row = template_row + 1
    clone_row(ws, template_row, new_row, {
        date_col: rday,
        reason_col: reason_text,
    })
    return True


def update_balance_sheet(ws, balance_items, mp_items, rday_str, tday_str):
    """잔고변경현황: 기존 동일 rday 블록 삭제 후 재작성"""
    rday = datetime.strptime(rday_str, '%Y-%m-%d')
    tday = datetime.strptime(tday_str, '%Y-%m-%d') if tday_str else rday

    # R4 헤더
    header_map = build_header_map(ws, 4)
    date_col = find_col(header_map, '리밸런싱일자')
    tday_col = find_col(header_map, '잔고변경일자')
    trade_type_col = find_col(header_map, '매매구분')
    isin_col = find_col(header_map, 'ISIN코드', 'ISIN 코드')
    qty_col = find_col(header_map, '보유수량')
    value_col = find_col(header_map, '평가금액')

    first_data_row = 5

    # MP 종목 전수 포함
    balance_isins = {i['isincode'] for i in balance_items}
    merged = list(balance_items)
    for mp in mp_items:
        if mp['isincode'] not in balance_isins:
            merged.append({
                'isincode': mp['isincode'],
                'trade_type': '변동없음',
                'qty': 0, 'value': 0,
            })
    # isincode가 None인 항목 제거
    merged = [m for m in merged if m.get('isincode')]
    merged.sort(key=lambda x: (x['isincode'] == 'CASH', x['isincode']))

    # 템플릿 정보 (삭제 전)
    tpl_data, tpl_sum, old_start, old_end = get_template_info(ws, date_col, isin_col)
    if not tpl_data:
        print('  [ERROR] 잔고변경현황에 템플릿 없음')
        return 0

    # 기존 rday 블록 삭제
    deleted = delete_block_with_sum(ws, date_col, rday, first_data_row, isin_col)
    if deleted:
        print(f'  기존 {deleted}행 삭제')
        tpl_data, tpl_sum, old_start, old_end = get_template_info(ws, date_col, isin_col)
        if not tpl_data:
            print('  [ERROR] 삭제 후 템플릿 없음')
            return 0

    # 새 블록 시작: 템플릿 합계행 바로 다음
    start_row = tpl_sum + 1

    for i, item in enumerate(merged):
        row = start_row + i
        clone_row(ws, tpl_data, row, {
            date_col: rday,
            tday_col: tday,
            trade_type_col: item.get('trade_type', ''),
            isin_col: item['isincode'],
            qty_col: item.get('qty', 0),
            value_col: item.get('value', 0),
        })

    # 합계행
    new_end = start_row + len(merged) - 1
    new_sum = new_end + 1
    clone_sum_row(ws, tpl_sum, new_sum,
                  old_start, old_end, start_row, new_end,
                  {date_col: rday, tday_col: tday, isin_col: '합계'})

    return len(merged)


# ===== 검증 =====

def collect_existing_isins(wb):
    ws = wb['투자유니버스']
    header_map = build_header_map(ws, 1)
    isin_col = find_col(header_map, '종목코드(ISIN코드)', 'ISIN코드', '종목코드')
    return {ws.cell(r, isin_col).value for r in range(2, ws.max_row + 1)
            if ws.cell(r, isin_col).value}


def validate(data, wb):
    errors, warnings, info = [], [], []
    rday = data['schedule']['rday']

    # MP 비중 합계
    for group, items in data['model_portfolio'].items():
        total = sum(i['ratio'] for i in items)
        if abs(total - 1.0) > 0.005:
            errors.append(f'MP({group}) 비중 합계 {total:.4f} — 허용범위 초과')
        elif abs(total - 1.0) > 0.0001:
            warnings.append(f'MP({group}) 비중 합계 {total:.6f} → 자동 보정')

    # 유니버스
    existing_isins = collect_existing_isins(wb)
    for isin in data['universe']:
        if isin != 'CASH' and isin not in existing_isins:
            warnings.append(f'유니버스 누락: {isin} → 자동 추가')

    if not data['trade_history']:
        warnings.append('거래내역 0건')

    # 중복 (경고만 — 삭제 후 재작성)
    rday_date = to_date(datetime.strptime(rday, '%Y-%m-%d'))
    for group in GROUPS:
        mp_ws = wb[sheet_names(group)['mp']]
        hdr = build_header_map(mp_ws, 1)
        dc = find_col(hdr, 'MP생성일자')
        for row in range(2, mp_ws.max_row + 1):
            if to_date(mp_ws.cell(row, dc).value) == rday_date:
                warnings.append(f'MP내역({group}): 기존 {rday} 데이터 → 삭제 후 재작성')
                break

    info.append(f'스케줄: {data["schedule"]["code"]} (rday={rday})')
    info.append(f'이전: {data["prev_schedule"]["code"]}')
    info.append(f'거래: {len(data["trade_history"])}건')
    for group in GROUPS:
        info.append(f'MP({group}): {len(data["model_portfolio"].get(group, []))}종목')

    return errors, warnings, info, existing_isins


# ===== 파일 관리 =====

def find_latest_file(recipe_dir, file_prefix):
    candidates = []
    for f in os.listdir(recipe_dir):
        if f.startswith(file_prefix) and '21.' in f and f.endswith('.xlsx') and '~$' not in f:
            m = re.search(r'(\d{8})\.xlsx$', f)
            if m:
                candidates.append((m.group(1), f))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def new_filename(old_filename, date_str):
    return re.sub(r'\d{8}\.xlsx$', f'{date_str}.xlsx', old_filename)


# ===== 메인 =====

def main():
    load_env()
    recipe_config = load_recipe_config()

    parser = argparse.ArgumentParser(description='서류21 리밸런싱 발생내역 엑셀 업데이트')
    parser.add_argument('--recipe', required=True, choices=list(recipe_config.keys()))
    parser.add_argument('--schedule-id', required=True, type=int)
    parser.add_argument('--dry-run', action='store_true', help='검증만, 파일 저장 안 함')
    args = parser.parse_args()

    config = recipe_config[args.recipe]
    base_dir = get_base_dir()
    recipe_dir = base_dir / config['dir']

    # 1. JSON 로드
    json_path = get_json_path(recipe_dir, args.schedule_id)
    if not json_path.exists():
        print(f'[ERROR] 데이터 없음: {json_path}')
        sys.exit(1)
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)
    print(f'데이터: {json_path}')

    # 2. 기존 파일
    latest_file = find_latest_file(recipe_dir, config['file_prefix'])
    if not latest_file:
        print(f'[ERROR] 서류21 파일 없음: {recipe_dir}')
        sys.exit(1)
    print(f'파일: {latest_file}')

    # 3. 엑셀 로드
    wb = openpyxl.load_workbook(recipe_dir / latest_file)
    print(f'시트: {wb.sheetnames}')

    # 4. 검증
    errors, warnings, info, existing_isins = validate(data, wb)
    print('\n=== 검증 ===')
    for i in info:
        print(f'  [INFO] {i}')
    for w in warnings:
        print(f'  [WARN] {w}')
    for e in errors:
        print(f'  [ERROR] {e}')
    if errors:
        sys.exit(1)
    if args.dry_run:
        print('\n--dry-run: 저장 안 함')
        return

    rday = data['schedule']['rday']

    # 5. 투자유니버스
    print('\n=== 투자유니버스 ===')
    added = update_universe(wb['투자유니버스'], data['universe'], existing_isins)
    print(f'  {len(added)}종목 추가' if added else '  변경 없음')

    # 6. 전체매매내역
    print('\n=== 전체매매내역 ===')
    count = update_trade_history(wb['전체매매내역'], data['trade_history'])
    print(f'  {count}건')

    # 7. 그룹별 시트
    pk_groups = data['pk_groups']
    for group in GROUPS:
        sheets = sheet_names(group)
        mp_items = data['model_portfolio'].get(group, [])
        print(f'\n=== {group} ===')

        n = update_mp(wb[sheets['mp']], mp_items, rday)
        print(f'  MP: {n}종목')

        ok = update_port_change(wb[sheets['port']], rday)
        print(f'  포트변경: {"완료" if ok else "실패"}')

        pks = pk_groups[group]
        for i, pk in enumerate(pks):
            bal = data['balance'].get(str(pk), [])
            tday = data['schedule']['tday']
            n = update_balance_sheet(wb[sheets['balance'][i]], bal, mp_items, rday, tday)
            print(f'  {sheets["balance"][i]}: {n}종목')

    # 8. 저장
    today_str = datetime.now().strftime('%Y%m%d')
    output_name = new_filename(latest_file, today_str)
    output_path = recipe_dir / output_name
    wb.save(output_path)
    print(f'\n=== 저장: {output_path} ===')


if __name__ == '__main__':
    main()
