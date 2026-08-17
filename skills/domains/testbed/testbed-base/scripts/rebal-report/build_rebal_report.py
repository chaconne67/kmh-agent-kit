# -*- coding: utf-8 -*-
"""3차+ 리밸런싱 발생내역(21번) DB기준 재구성 빌더.

구조: [운용개시 블록] + [리밸런싱 회차 블록]. 자산종류=기초이름+위험도점수(세분류).
자세한 절차/함정은 workflows/rebal-report/rebal-report.md 의 "3차 재구성 방식" 참조.

입력 JSON (런타임 생성 — 워크플로우의 조회 단계에서 만든다):
  existing_uni.json   기존 투자유니버스 (20번/직전 21번 시트에서 추출. 종목 보존용)
  uni_full.json       전략 후보 유니버스 (BacktestTicker 합집합) + DB분류 (ticker,isin,name,market,asset_type,dscore,is_danger)
  extra_db.json       기존 유니버스 중 후보에 없는 종목의 DB분류 (uni_full와 동일 스키마)
  opening_block.json  운용개시 블록 (groups[g].mp/mp_date/limits, balance[acct].rday/bday/items)  ★잔고는 5행부터 추출
  sheet_meta.json     balance_hdr[acct]{유형,계좌번호,운용금액}, pc_param[g]{유형,위험자산한도,최저위험도,최고위험도,단일종목한도}
  rebalancing_data_{ID}.json  fetch_rebalancing_data.py 산출 (schedule,model_portfolio,balance,pk_groups,trade_history)

CONFIG는 매 회차/레시피마다 사용자가 검토·수정한다(런타임 값).
"""
import json, re, shutil, sys
from copy import copy
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter as L

# ===================== CONFIG (런타임 값 — 매번 검토) =====================
WORK_DIR   = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()  # JSON 입력 + 출력 디렉토리
TEMPLATE   = Path(sys.argv[2])  # Drive에서 받은 로컬 21번 xlsx 서식/구조 베이스
OUT        = Path(sys.argv[3])  # 출력 xlsx 경로
FETCH_JSON = WORK_DIR / (sys.argv[4] if len(sys.argv) > 4 else 'rebalancing_data.json')
RDAY       = sys.argv[5] if len(sys.argv) > 5 else None  # 회차 리밸런싱일자(기준일). 스케줄 rday와 다를 수 있음(사용자 통일)
REASON     = sys.argv[6] if len(sys.argv) > 6 else '정기리밸런싱'

# DB asset_type 오분류/기초이름 보정 (★매 실행 검토: DB가 틀린 종목만. 라이브 영향 피해 DB 직접수정 대신 보고서 보정)
#   WORK_DIR/asset_type_fix.json 에 {ticker: 보정된_기초이름(숫자 suffix 없이)} 형태로 둔다. 없으면 빈 dict.
#   예: {"275300":"국내주식", "423160":"국내안전자산채권"}  (KODEX우량주 해외→국내 오분류, KODEX KOFR bare→국내안전자산채권)
# ========================================================================

existing = json.loads((WORK_DIR/'existing_uni.json').read_text(encoding='utf-8'))
btrows   = json.loads((WORK_DIR/'uni_full.json').read_text(encoding='utf-8'))
extra    = json.loads((WORK_DIR/'extra_db.json').read_text(encoding='utf-8'))
opening  = json.loads((WORK_DIR/'opening_block.json').read_text(encoding='utf-8'))
meta     = json.loads((WORK_DIR/'sheet_meta.json').read_text(encoding='utf-8'))
fetch    = json.loads(FETCH_JSON.read_text(encoding='utf-8'))
_atfix   = WORK_DIR/'asset_type_fix.json'
ASSET_TYPE_FIX = json.loads(_atfix.read_text(encoding='utf-8')) if _atfix.exists() else {}
if RDAY is None:
    RDAY = fetch['schedule']['rday']

DB = {r['ticker']: r for r in (btrows + extra)}
GRADE_TEXT = {1:'매우낮은위험',2:'낮은위험',3:'보통위험',4:'다소높은위험',5:'높은위험',6:'매우높은위험'}
GROUPS = ['안정','중립','적극']
NEXT   = {'안정':'중립','중립':'적극','적극':'중립'}  # 차별성 비교대상

def is_domestic(r):
    return r['ticker'] == 'CASH' or r.get('market') in ('KRX','KOSPI','KOSDAQ')
def cbase(r):
    b = re.sub(r'\d+$', '', str(r['asset_type']))
    return ASSET_TYPE_FIX.get(r['ticker'], b)
def rep_at(r):  # 보고서 자산종류 = 기초이름 + 위험도점수(dscore). 현금은 숫자 없이.
    return '현금' if r['ticker'] == 'CASH' else cbase(r) + str(r['dscore'])

# ===== 투자유니버스: 기존 보존(자산종류 DB갱신) + 국내후보 누락분 추가. 해외상장 제외 =====
def build_universe(ws):
    styles = [copy(ws.cell(2,c)._style) for c in range(1,10)]
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row-1)
    final, ex_tickers = [], {e['ticker'] for e in existing}
    for e in existing:                       # 기존 종목 전부 보존
        row = dict(e); d = DB.get(e['ticker'])
        if d:
            row['asset_type']=rep_at(d); row['dscore']=d['dscore']
            row['grade']=GRADE_TEXT.get(d['dscore'],e['grade']); row['danger']='Y' if d['is_danger'] else 'N'
        elif e['isin'] in (None,'None') or (e['name'] and '현금' in str(e['name'])) or e['ticker'] in ('None','없음'):
            row['asset_type']='현금'
        final.append(row)
    for d in btrows:                          # 국내 전략후보 중 기존에 없는 것 추가
        if is_domestic(d) and d['ticker'] not in ex_tickers and d['ticker']!='CASH':
            final.append({'isin':d['isin'],'name':d['name'],'sigu':'국내','jasan_gun':'ETF',
                          'asset_type':rep_at(d),'grade':GRADE_TEXT.get(d['dscore'],''),
                          'dscore':d['dscore'],'danger':'Y' if d['is_danger'] else 'N','ticker':d['ticker']})
    for i,row in enumerate(final):
        for c,v in enumerate([row['isin'],row['name'],row['sigu'],row['jasan_gun'],row['asset_type'],
                              row['grade'],row['dscore'],row['danger'],row['ticker']], 1):
            cell=ws.cell(2+i,c,v)
            try: cell._style=styles[c-1]
            except: pass
    at_score={row['asset_type']: row['dscore'] for row in final}
    return at_score

def make_header_order(at_score):  # 자산종류 헤더 순서: 위험도점수 desc, 현금 마지막
    ats=[a for a in at_score if a!='현금']
    ats.sort(key=lambda a:(-at_score[a], a))
    return ats + (['현금'] if '현금' in at_score else [])

# ===== MP내역 (자산종류 D = 투자유니버스 E 참조 수식) =====
def f_C(r): return f'=IF(OR(B{r}="",B{r}="합계"),"",INDEX(투자유니버스!B:B,MATCH($B{r},투자유니버스!$A:$A,0)))'
def f_D(r): return f'=IF(OR(B{r}="",B{r}="합계"),"",INDEX(투자유니버스!E:E,MATCH($B{r},투자유니버스!$A:$A,0)))'
def f_E(r): return f'=IF(OR(B{r}="",B{r}="합계"),"",INDEX(투자유니버스!G:G,MATCH($B{r},투자유니버스!$A:$A,0)))'
def f_F(r): return f'=IF(OR(B{r}="",B{r}="합계"),"",INDEX(투자유니버스!H:H,MATCH($B{r},투자유니버스!$A:$A,0)))'
def f_H(r): return f'=IF(OR(B{r}="",B{r}="합계"),"",IF(COUNTIF(투자유니버스!A:A,B{r})>0,"O","X"))'

def mp_opening(g):
    return [{'isin':x['isin'],'ratio':x['ratio']} for x in opening['groups'][g]['mp']], opening['groups'][g]['mp_date']
def mp_rebal(g):
    items=[{'isin':('CASH' if x['isincode']=='CASH' else x['isincode']),'ratio':round(float(x['ratio']),6)}
           for x in fetch['model_portfolio'][g]]
    tot=sum(i['ratio'] for i in items)           # 비중합 100% 보정 (±0.5% 이내 CASH/마지막 잔차)
    if 0.0001 < abs(tot-1.0) <= 0.005:
        tgt=next((i for i in items if i['isin']=='CASH'), items[-1]); tgt['ratio']=round(tgt['ratio']+round(1.0-tot,6),6)
    return items, RDAY

def sort_cash_last(items, key):
    non=[i for i in items if key(i)!='CASH']; cash=[i for i in items if key(i)=='CASH']
    non.sort(key=lambda x:str(key(x))); return non+cash

def build_mp(ws, g):
    st_d=[copy(ws.cell(2,c)._style) for c in range(1,9)]
    sumr=next((r for r in range(2,ws.max_row+1) if ws.cell(r,2).value=='합계'), None)
    st_s=[copy(ws.cell(sumr,c)._style) for c in range(1,9)] if sumr else st_d
    if ws.max_row>=2: ws.delete_rows(2, ws.max_row-1)
    r=2
    for items,date in [mp_opening(g), mp_rebal(g)]:
        items=sort_cash_last(items, lambda x:x['isin']); d=datetime.strptime(date,'%Y-%m-%d'); start=r
        for it in items:
            ws.cell(r,1,d); ws.cell(r,2,it['isin']); ws.cell(r,3,f_C(r)); ws.cell(r,4,f_D(r))
            ws.cell(r,5,f_E(r)); ws.cell(r,6,f_F(r)); ws.cell(r,7,round(float(it['ratio']),6)); ws.cell(r,8,f_H(r))
            for c in range(1,9):
                try: ws.cell(r,c)._style=st_d[c-1]
                except: pass
            r+=1
        ws.cell(r,1,d); ws.cell(r,2,'합계'); ws.cell(r,3,f_C(r)); ws.cell(r,4,f_D(r))
        ws.cell(r,5,f_E(r)); ws.cell(r,6,f_F(r)); ws.cell(r,7,f'=SUM(G{start}:G{r-1})'); ws.cell(r,8,f_H(r))
        for c in range(1,9):
            try: ws.cell(r,c)._style=st_s[c-1]
            except: pass
        r+=1

# ===== 포트변경내역 (자산종류 세분류 헤더 N개) =====
def build_port(ws, g, header_order, at_score):
    mp=f"'MP내역({g})'"; pcg=f"'포트변경내역({g})'"; pcn=f"'포트변경내역({NEXT[g]})'"
    p=meta['pc_param'][g]; n=len(header_order)
    # 템플릿 자산컬럼 수(B..합계직전) 탐지 후 N개로 맞춰 잉여 컬럼 삭제. 헤더 서식/병합 보존.
    sum_col=next(c for c in range(2,ws.max_column+1) if str(ws.cell(4,c).value)=='합계')
    tpl_asset_n=sum_col-2  # B..(sum_col-1)
    for row in ws.iter_rows():               # 코멘트 제거(병합셀 코멘트가 delete 충돌)
        for cell in row:
            if cell.comment is not None: cell.comment=None
    for m in list(ws.merged_cells.ranges):   # 병합 해제(delete/insert가 병합 갱신 못함)
        ws.unmerge_cells(str(m))
    # 템플릿 자산컬럼 수를 N에 맞춤: 많으면 삭제, 적으면 삽입(스타일/너비 복사). 합계 이후는 자동 시프트.
    if tpl_asset_n > n:
        ws.delete_cols(2+n, tpl_asset_n-n)
    elif tpl_asset_n < n:
        add=n-tpl_asset_n
        ws.insert_cols(2+tpl_asset_n, add)
        for off in range(add):                # 삽입 컬럼에 자산컬럼(B) 스타일/너비 복사
            cidx=2+tpl_asset_n+off
            ws.column_dimensions[L(cidx)].width = ws.column_dimensions[L(2)].width
            for rr in range(1, 7):
                try: ws.cell(rr,cidx)._style = copy(ws.cell(rr,2)._style)
                except: pass
    # 컬럼 위치(자산 N개 후): 합계=2+n, 위험자산비중=+1, 위험도=+2, 사유=+3, 검증 6개=+4..+9
    Csum,Crisk,Crd,Crsn=2+n,3+n,4+n,5+n
    Cv1,Cv2,Cv3,Cv4,Cv5,Cv6=6+n,7+n,8+n,9+n,10+n,11+n
    ws.merge_cells(f'{L(Csum)}4:{L(Csum)}5'); ws.merge_cells(f'{L(Crisk)}4:{L(Crisk)}5')
    ws.merge_cells(f'{L(Crd)}4:{L(Crd)}5'); ws.merge_cells(f'{L(Crsn)}4:{L(Crsn)}5')
    ws.merge_cells(f'{L(Cv1)}4:{L(Cv3)}4')      # 포트폴리오 변경내역 확인 (위험자산비중적정성/위험도적정성/차별성)
    ws.merge_cells(f'{L(Cv4)}4:{L(Cv6+1)}4')    # 잔고변경현황 확인 (종목수/단일종목한도/자료정합성/투자유니버스) 4칸
    ws.merge_cells(f'{L(Cv6+2)}4:{L(Cv6+2)}5')  # 부적합 사유
    data_style=[copy(ws.cell(6,c)._style) for c in range(1,Cv6+2)]
    if ws.max_row>=6: ws.delete_rows(6, ws.max_row-5)
    ws.cell(2,1,p['유형']); ws.cell(2,2,'국내ETF레시피')
    ws.cell(2,3,p['위험자산한도']); ws.cell(2,4,p['최저위험도']); ws.cell(2,5,p['최고위험도']); ws.cell(2,6,p['단일종목한도'])
    for idx,at in enumerate(header_order):
        ws.cell(4,2+idx,at); ws.cell(5,2+idx,at_score[at])
    last_asset=L(1+n)  # 마지막 자산컬럼 (B=2 → 1+n)
    blocks=[(opening['groups'][g]['mp_date'],'운용개시'), (RDAY,REASON)]
    r=6
    for date,reason in blocks:
        d=datetime.strptime(date,'%Y-%m-%d'); ws.cell(r,1,d)
        for idx,at in enumerate(header_order):
            ws.cell(r,2+idx, f'=IF(ISBLANK($A{r}),"",SUMIFS({mp}!$G:$G,{mp}!$A:$A,$A{r},{mp}!$D:$D,{L(2+idx)}$4))')
        ws.cell(r,Csum, f'=IF(ISBLANK(A{r}),"",SUM(B{r}:{last_asset}{r}))')
        ws.cell(r,Crisk,f'=IF(ISBLANK(A{r}),"",SUMIFS({mp}!G:G,{mp}!A:A,{pcg}!A{r},{mp}!F:F,"Y"))')
        ws.cell(r,Crd,  f'=IF(ISBLANK(A{r}),"",SUMPRODUCT($B$5:${last_asset}$5,B{r}:{last_asset}{r}))')
        ws.cell(r,Crsn, reason)
        cr=L(Crd)
        ws.cell(r,Cv1, f'=IF(ISBLANK(A{r}),"",IF($C$2>={L(Crisk)}{r},"O","X"))')
        ws.cell(r,Cv2, f'=IF(ISBLANK(A{r}),"",IF(AND($D$2<={cr}{r},{cr}{r}<=$E$2),"O","X"))')
        opx='<' if g!='적극' else '>'
        ws.cell(r,Cv3, f'=IF(ISBLANK(A{r}),"",IFERROR(IF({cr}{r}{opx}INDEX({pcn}!{cr}:{cr},MATCH(A{r},{pcn}!A:A,0)),"O","X"),""))')
        ws.cell(r,Cv4, f'=IF(ISBLANK(A{r}),"",COUNTIFS({mp}!$A:$A,A{r},{mp}!G:G,">0",{mp}!$B:$B,"<>현금",{mp}!$B:$B,"<>예수금",{mp}!$B:$B,"<>예탁금",{mp}!$B:$B,"<>합계"))')
        ws.cell(r,Cv5, f'=IF(ISBLANK(A{r}),"",IF(COUNTIFS({mp}!A:A,A{r},{mp}!G:G,">"&$F$2,{mp}!B:B,"<>현금",{mp}!B:B,"<>합계")=0,"O","X"))')
        ws.cell(r,Cv6, f'=IF(ISBLANK(A{r}),"",IF(AND(ABS({L(Crisk)}{r}-SUMIFS({mp}!G:G,{mp}!A:A,A{r},{mp}!F:F,"Y"))<0.001,ABS({L(Csum)}{r}-SUMIFS({mp}!G:G,{mp}!A:A,A{r},{mp}!B:B,"<>합계"))<0.001),"O","X"))')
        ws.cell(r,Cv6+1, f'=IF(ISBLANK(A{r}),"",IF(COUNTIFS({mp}!A:A,A{r},{mp}!H:H,"X")=0,"O","X"))')
        for c in range(1,Cv6+2):
            try: ws.cell(r,c)._style=data_style[c-1]
            except: pass
        r+=1

# ===== 잔고변경현황 (★데이터 5행부터. MP목표종목 전부 포함 → 목표비중 100%) =====
def bf_E(r): return f'=IF(OR(D{r}="",D{r}="합계"),"",INDEX(투자유니버스!B:B,MATCH($D{r},투자유니버스!$A:$A,0)))'
def bf_F(r): return f'=IF(OR(D{r}="",D{r}="합계"),"",INDEX(투자유니버스!E:E,MATCH($D{r},투자유니버스!$A:$A,0)))'

def build_balance(ws, g, idx, acctkey):
    mp=f"'MP내역({g})'"; hdr=meta['balance_hdr'][acctkey]
    pk=fetch['pk_groups'][g][idx]; r33items=fetch['balance'].get(str(pk),[])
    op=opening['balance'][acctkey]; opq={it['isin']:it['qty'] for it in op['items']}
    data_style=[copy(ws.cell(5,c)._style) for c in range(1,13)]
    sumr=next((r for r in range(5,ws.max_row+1) if any(str(ws.cell(r,c).value)=='합계' for c in range(1,13))), None)
    sum_style=[copy(ws.cell(sumr,c)._style) for c in range(1,13)] if sumr else data_style
    if ws.max_row>=5: ws.delete_rows(5, ws.max_row-4)  # ★5행부터 (헤더 4행 보존)
    ws.cell(2,1,hdr['유형']); ws.cell(2,2,'국내ETF레시피'); ws.cell(2,3,hdr['계좌번호']); ws.cell(2,4,hdr['운용금액'])
    op_mp=[x['isin'] for x in opening['groups'][g]['mp']]
    r33_mp=[x['isincode'] for x in fetch['model_portfolio'][g]]
    blocks=[]
    op_block=[]; seen=set()
    for it in op['items']: op_block.append((it['isin'],it['매매구분'],it['qty'],it['value'])); seen.add(it['isin'])
    for isin in op_mp:
        if isin not in seen: op_block.append((isin,'변동없음',0,0)); seen.add(isin)
    blocks.append((op['rday'], op['bday'], op_block))
    r33=[]; seen33=set()
    for it in r33items:
        isin=it['isincode']; seen33.add(isin)
        if isin=='CASH': r33.append((isin,'변동없음',it.get('qty',0),it.get('value',0))); continue
        nq=it.get('qty',0); pq=opq.get(isin,0)
        gb=('신규매수' if pq==0 and nq>0 else '추가매수' if nq>pq else '일부매도' if 0<nq<pq else '전량매도' if nq==0 and pq>0 else '변동없음')
        r33.append((isin,gb,nq,it.get('value',0)))
    for isin in r33_mp:
        if isin not in seen33: r33.append((isin, '전량매도' if opq.get(isin,0)>0 else '변동없음', 0, 0)); seen33.add(isin)
    blocks.append((RDAY, fetch['schedule']['tday'], r33))
    r=5
    for rday,bday,items in blocks:
        rd=datetime.strptime(rday,'%Y-%m-%d'); bd=datetime.strptime(bday,'%Y-%m-%d'); start=r
        items=sort_cash_last(items, lambda x:x[0])
        for isin,gb,qty,val in items:
            ws.cell(r,1,rd); ws.cell(r,2,bd); ws.cell(r,3,gb); ws.cell(r,4,isin)
            ws.cell(r,5,bf_E(r)); ws.cell(r,6,bf_F(r)); ws.cell(r,7,qty); ws.cell(r,8,round(float(val),2) if val else 0)
            ws.cell(r,9,f'=H{r}/SUMIF(B:B,B{r},H:H)'); ws.cell(r,10,f"=SUMIFS({mp}!G:G,{mp}!A:A,A{r},{mp}!B:B,D{r})"); ws.cell(r,11,f'=ABS(I{r}-J{r})')
            for c in range(1,13):
                try: ws.cell(r,c)._style=data_style[c-1]
                except: pass
            r+=1
        ws.cell(r,1,rd); ws.cell(r,2,bd); ws.cell(r,4,'합계')
        ws.cell(r,9,f'=SUM(I{start}:I{r-1})'); ws.cell(r,10,f'=SUM(J{start}:J{r-1})'); ws.cell(r,11,f'=SUM(K{start}:K{r-1})')
        for c in range(1,13):
            try: ws.cell(r,c)._style=sum_style[c-1]
            except: pass
        r+=1

# ===== 전체매매내역 (헤더 동적 매핑, 레시피명 컬럼 = 미매핑 빈헤더) =====
def build_trades(ws):
    hmap={str(ws.cell(1,c).value).strip():c for c in range(1,ws.max_column+1) if ws.cell(1,c).value is not None}
    def col(*names):
        for n in names:
            if n in hmap: return hmap[n]
        for n in names:
            for k,cc in hmap.items():
                if n in k: return cc
    cd,ct,ca=col('매매일자'),col('포트폴리오유형'),col('일임계좌번호')
    cg,cn,cq,cp,cb=col('매매구분'),col('종목명'),col('매매수량'),col('매매가격'),col('잔고수량')
    cc=col('종목코드','ISIN코드')
    mapped={cd,ct,ca,cg,cn,cq,cp,cb,cc}
    crec=next((c for c in range(1,ws.max_column+1) if ws.cell(1,c).value is not None and c not in mapped), None)
    data_style=[copy(ws.cell(2,c)._style) for c in range(1,ws.max_column+1)]
    if ws.max_row>=2: ws.delete_rows(2, ws.max_row-1)
    r=2
    for t in fetch['trade_history']:
        for c,v in [(cd,t['매매일자'] and datetime.strptime(t['매매일자'],'%Y-%m-%d')),(ct,t.get('포트폴리오유형')),
                    (crec,t.get('모아에셋포트명')),(ca,t.get('일임계좌번호')),(cg,t.get('매매구분')),
                    (cc,t.get('종목코드')),(cn,t.get('종목명')),(cq,t.get('매매수량')),(cp,t.get('매매가격')),(cb,t.get('잔고수량'))]:
            if c: ws.cell(r,c,v)
        for c in range(1,ws.max_column+1):
            try: ws.cell(r,c)._style=data_style[c-1]
            except: pass
        r+=1

def fix_guide(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value,str) and '20.포트폴리오 운용개시 현황' in cell.value:
                cell.value=cell.value.replace('20.포트폴리오 운용개시 현황','21.리밸런싱 발생내역')

# ===== 실행 =====
shutil.copy(TEMPLATE, OUT)
wb = openpyxl.load_workbook(OUT)
at_score = build_universe(wb['투자유니버스'])
header_order = make_header_order(at_score)
for g in GROUPS: build_mp(wb[f'MP내역({g})'], g)
for g in GROUPS: build_port(wb[f'포트변경내역({g})'], g, header_order, at_score)
for g in GROUPS:
    for i in range(3): build_balance(wb[f'잔고변경현황({g}{i+1})'], g, i, f'{g}{i+1}')
build_trades(wb['전체매매내역'])
fix_guide(wb['작성가이드'])
wb.save(OUT)
print('빌드 완료:', OUT)
print('자산종류 헤더:', header_order)
