# -*- coding: utf-8 -*-
"""20번 운용개시 현황 xlsx → 빌더 입력 JSON 생성 (existing_uni / opening_block / sheet_meta).
사용: python prepare_local_data.py "<20번_xlsx_경로>" "<출력_WORK_DIR>"

★ 함정:
  - 잔고변경현황 데이터는 **5행부터** 시작(헤더 4행). 6행부터 읽으면 첫 종목 누락.
  - MP내역 데이터는 2행부터(헤더 1행). 투자유니버스도 2행부터.
  - 한글 경로는 pathlib로만 처리(Bash ls/cp 금지).
  - Google Drive 연결 기능으로 받은 raw 20번 xlsx가 실제로 열리는지 확인한 뒤 사용.
"""
import json, sys
from pathlib import Path
from datetime import datetime, date
import openpyxl

F20 = Path(sys.argv[1]); WORK = Path(sys.argv[2]); WORK.mkdir(parents=True, exist_ok=True)
wb = openpyxl.load_workbook(F20)
def ds(v): return v.strftime('%Y-%m-%d') if isinstance(v,(datetime,date)) else v

# 투자유니버스 (기존 종목 보존용) — 컬럼: A ISIN, B 종목명, C 시장, D 자산군, E 자산종류, F 위험등급, G 위험도점수, H 위험자산여부, I 비고(ticker)
uni=wb['투자유니버스']; existing=[]
for r in range(2, uni.max_row+1):
    a=uni.cell(r,1).value
    if not a: continue
    existing.append({'isin':str(a),'name':uni.cell(r,2).value,'sigu':uni.cell(r,3).value,
                     'jasan_gun':uni.cell(r,4).value,'asset_type':uni.cell(r,5).value,
                     'grade':uni.cell(r,6).value,'dscore':uni.cell(r,7).value,
                     'danger':uni.cell(r,8).value,'ticker':str(uni.cell(r,9).value)})
(WORK/'existing_uni.json').write_text(json.dumps(existing,ensure_ascii=False,indent=1),encoding='utf-8')

# 운용개시 블록: groups[g].mp(2행~)/mp_date/limits, balance[acct](5행~)
out={'groups':{}, 'balance':{}}
meta={'balance_hdr':{}, 'pc_param':{}}
for g in ['안정','중립','적극']:
    mp=wb[f'MP내역({g})']; rows=[]; mpdate=None
    for r in range(2, mp.max_row+1):
        b=mp.cell(r,2).value
        if not b or b=='합계': continue
        rows.append({'isin':str(b),'ratio':mp.cell(r,7).value})
        if mpdate is None: mpdate=ds(mp.cell(r,1).value)
    pc=wb[f'포트변경내역({g})']
    out['groups'][g]={'mp_date':mpdate,'mp':rows,
                      'limits':{'위험자산한도':pc.cell(2,3).value,'최저위험도':pc.cell(2,4).value,
                                '최고위험도':pc.cell(2,5).value,'단일종목한도':pc.cell(2,6).value}}
    meta['pc_param'][g]={'유형':pc.cell(2,1).value,'위험자산한도':pc.cell(2,3).value,'최저위험도':pc.cell(2,4).value,
                         '최고위험도':pc.cell(2,5).value,'단일종목한도':pc.cell(2,6).value}
for g in ['안정','중립','적극']:
    for i in [1,2,3]:
        ws=wb[f'잔고변경현황({g}{i})']; items=[]; rday=bday=None
        for r in range(5, ws.max_row+1):       # ★ 5행부터
            d=ws.cell(r,4).value
            if not d or d=='합계': continue
            items.append({'isin':str(d),'매매구분':ws.cell(r,3).value,'qty':ws.cell(r,7).value,'value':ws.cell(r,8).value})
            if rday is None: rday=ds(ws.cell(r,1).value); bday=ds(ws.cell(r,2).value)
        out['balance'][f'{g}{i}']={'rday':rday,'bday':bday,'items':items}
        meta['balance_hdr'][f'{g}{i}']={'유형':ws.cell(2,1).value,'계좌번호':ws.cell(2,3).value,'운용금액':ws.cell(2,4).value}
(WORK/'opening_block.json').write_text(json.dumps(out,ensure_ascii=False,indent=1),encoding='utf-8')
(WORK/'sheet_meta.json').write_text(json.dumps(meta,ensure_ascii=False,indent=1),encoding='utf-8')
print('생성:', [p for p in ['existing_uni.json','opening_block.json','sheet_meta.json']])
print('기존 유니버스', len(existing), '종목 / 운용개시 MP', {g:len(out["groups"][g]["mp"]) for g in out["groups"]})
