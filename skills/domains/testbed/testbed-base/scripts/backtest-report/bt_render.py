"""Pure XLSX renderer. Reads single JSON, creates xlsx from zero base."""
import argparse, json, pathlib, sys
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, AreaChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color

sys.stdout.reconfigure(encoding='utf-8')


def main():
    parser = argparse.ArgumentParser(description='Render xlsx from JSON')
    parser.add_argument('--input', required=True, help='bt_render_input.json path')
    parser.add_argument('--output', required=True, help='Output xlsx path')
    parser.add_argument('--ref', help='Reference xlsx for direct style copy (bypasses JSON styles)')
    parser.add_argument('--spec', help='bt_spec.json for region info (required with --ref)')
    args = parser.parse_args()
    data = json.loads(pathlib.Path(args.input).read_text(encoding='utf-8'))
    ref_path = pathlib.Path(args.ref) if args.ref else None
    spec = json.loads(pathlib.Path(args.spec).read_text(encoding='utf-8')) if args.spec else None
    render(data, pathlib.Path(args.output), ref_path=ref_path, spec=spec)


def render(data, output_path, ref_path=None, spec=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_data in data['sheets']:
        ws = wb.create_sheet(title=sheet_data['name'])
        render_sheet(ws, sheet_data)
    # Style override pass: copy styles directly from reference (no JSON round-trip)
    if ref_path and spec:
        _apply_ref_styles(wb, ref_path, spec, data)
    wb.save(output_path)
    print(f'Saved: {output_path}')


def json_to_font(d):
    if not d:
        return Font()
    color = d.get('color')
    font_color = None
    if isinstance(color, str) and color.startswith('#'):
        font_color = color[1:]
    elif isinstance(color, dict):
        if 'theme' in color:
            font_color = Color(theme=color['theme'], tint=color.get('tint', 0))
        elif 'indexed' in color:
            font_color = Color(indexed=color['indexed'])
    return Font(
        name=d.get('name'),
        size=d.get('size', 11),
        bold=d.get('bold', False),
        italic=d.get('italic', False),
        underline=d.get('underline'),
        strikethrough=d.get('strikethrough', False),
        color=font_color,
    )


def json_to_fill(d):
    if not d or not d.get('type'):
        return PatternFill()
    fg = d.get('fgColor')
    fg_color = None
    if isinstance(fg, str) and fg.startswith('#'):
        fg_color = fg[1:]
    elif isinstance(fg, dict):
        if 'theme' in fg:
            fg_color = Color(theme=fg['theme'], tint=fg.get('tint', 0))
        elif 'indexed' in fg:
            fg_color = Color(indexed=fg['indexed'])
    return PatternFill(fill_type=d['type'], fgColor=fg_color)


def json_to_border(d):
    if not d:
        return Border()
    def make_side(sd):
        if not sd:
            return Side()
        color = sd.get('color')
        c = None
        if isinstance(color, str) and color.startswith('#'):
            c = color[1:]
        elif isinstance(color, dict) and 'theme' in color:
            c = Color(theme=color['theme'], tint=color.get('tint', 0))
        return Side(style=sd.get('style'), color=c)
    return Border(
        left=make_side(d.get('left')),
        right=make_side(d.get('right')),
        top=make_side(d.get('top')),
        bottom=make_side(d.get('bottom')),
    )


def json_to_alignment(d):
    if not d:
        return Alignment()
    return Alignment(
        horizontal=d.get('horizontal'),
        vertical=d.get('vertical'),
        wrap_text=d.get('wrap_text'),
        text_rotation=d.get('text_rotation'),
    )


def apply_style(cell, style_dict):
    if not style_dict:
        return
    cell.font = json_to_font(style_dict.get('font'))
    cell.fill = json_to_fill(style_dict.get('fill'))
    cell.border = json_to_border(style_dict.get('border'))
    cell.alignment = json_to_alignment(style_dict.get('alignment'))
    if style_dict.get('number_format'):
        cell.number_format = style_dict['number_format']


def render_sheet(ws, sheet_data):
    # Column widths
    for col_letter, width in sheet_data.get('column_widths', {}).items():
        ws.column_dimensions[col_letter].width = width
    # Row heights
    for row_str, height in sheet_data.get('row_heights', {}).items():
        ws.row_dimensions[int(row_str)].height = height
    # Cells
    for row_data in sheet_data.get('rows', []):
        r = row_data['row_num']
        for cell_data in row_data.get('cells', []):
            c = cell_data['col']
            val = cell_data.get('value')
            # Handle date values
            if cell_data.get('value_type') == 'date' and isinstance(val, str):
                try:
                    val = datetime.strptime(val[:10], '%Y-%m-%d')
                except ValueError:
                    pass
            ws.cell(r, c, val)
            apply_style(ws.cell(r, c), cell_data.get('style'))
    # Merged cells (after writing values)
    for merge_range in sheet_data.get('merged_cells', []):
        ws.merge_cells(merge_range)
    # Charts
    for chart_data in sheet_data.get('charts', []):
        render_chart(ws, chart_data)


def _restore_chart_from_xml(chart_xml):
    """Restore a chart object from full chartSpace XML."""
    from lxml import etree
    from openpyxl.chart.chartspace import ChartSpace
    from openpyxl.chart.reader import read_chart
    try:
        tree = etree.fromstring(chart_xml)
        cs = ChartSpace.from_tree(tree)
        return read_chart(cs)
    except Exception as e:
        print(f'Warning: chart XML restore failed: {e}')
        return None


def _build_chart_from_props(chart_data):
    """Fallback: build chart from individual properties (legacy mode)."""
    chart_type = chart_data.get('type', 'LineChart')
    chart_classes = {'LineChart': LineChart, 'BarChart': BarChart, 'AreaChart': AreaChart}
    ChartClass = chart_classes.get(chart_type, LineChart)
    chart = ChartClass()
    if chart_data.get('grouping'):
        chart.grouping = chart_data['grouping']
    if chart_data.get('style') is not None:
        chart.style = chart_data['style']
    if chart_data.get('width'):
        chart.width = chart_data['width']
    if chart_data.get('height'):
        chart.height = chart_data['height']
    return chart


def _replace_data_refs(chart, ws, chart_data):
    """Replace chart data references with target ranges, then add to worksheet."""
    # Clear existing series (from XML restore) and re-add with target refs
    per_series = chart_data.get('per_series_refs')
    dr = chart_data.get('data_ref', {})
    cr = chart_data.get('cat_ref', {})

    if per_series or dr:
        # Remove old series from XML restore
        chart.series = []

        if per_series:
            for sr in per_series:
                val_ref = Reference(ws, min_col=sr['val_col'], min_row=1,
                                    max_row=dr.get('max_row', 100))
                chart.add_data(val_ref, titles_from_data=chart_data.get('titles_from_data', True))
        elif dr:
            data_ref = Reference(ws, min_col=dr['min_col'],
                                 max_col=dr.get('max_col', dr['min_col']),
                                 min_row=dr['min_row'], max_row=dr['max_row'])
            chart.add_data(data_ref, titles_from_data=chart_data.get('titles_from_data', True))

        if cr:
            cat_ref = Reference(ws, min_col=cr['min_col'], min_row=cr['min_row'],
                                max_row=cr['max_row'])
            chart.set_categories(cat_ref)

    # Re-apply series styling from spec (data refs changed, styles need reattach)
    spec_series = chart_data.get('series', [])
    for i, s in enumerate(chart.series):
        if i < len(spec_series):
            ss = spec_series[i]
            if 'fill_color' in ss:
                apply_chart_color(s, ss['fill_color'], 'fill')
            if 'line_color' in ss:
                apply_chart_color(s, ss['line_color'], 'line')
            if 'line_width' in ss:
                from openpyxl.drawing.line import LineProperties as LP
                if s.graphicalProperties.line is None:
                    s.graphicalProperties.line = LP()
                s.graphicalProperties.line.width = ss['line_width']

    # Add chart to worksheet with anchor
    anchor_str = chart_data.get('anchor', 'A1')
    ws.add_chart(chart, anchor_str)

    # Apply precise anchor offset
    ad = chart_data.get('ref_anchor_detail')
    if ad and (ad.get('colOff', 0) or ad.get('rowOff', 0)):
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
        w_emu = int(chart.width * 914400)
        h_emu = int(chart.height * 914400)
        from_marker = AnchorMarker(col=ad['col'], colOff=ad.get('colOff', 0),
                                   row=ad['row'], rowOff=ad.get('rowOff', 0))
        to_col = ad['col'] + max(1, w_emu // 914400)
        to_row = ad['row'] + max(1, h_emu // 200000)
        to_marker = AnchorMarker(col=to_col, row=to_row)
        chart.anchor = TwoCellAnchor(_from=from_marker, to=to_marker)


def render_chart(ws, chart_data):
    """Render chart from spec. Two modes:
    1. Full XML restore (_chart_xml present): rebuild chart from XML, replace data refs
    2. Fallback: build chart from individual properties
    """
    chart_xml = chart_data.get('_chart_xml')
    if chart_xml:
        chart = _restore_chart_from_xml(chart_xml)
    else:
        chart = _build_chart_from_props(chart_data)

    if chart is None:
        return

    # Replace data references and add to worksheet (includes series styling + anchor)
    _replace_data_refs(chart, ws, chart_data)


def _apply_axis_line(axis, line_spec):
    """Apply axis line formatting (color, visibility)."""
    if not line_spec:
        return
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    ln = LineProperties(w=line_spec.get('width'), prstDash=line_spec.get('prstDash'))
    if line_spec.get('noFill'):
        ln.noFill = True
    elif line_spec.get('solidFill'):
        apply_chart_line_color(ln, line_spec['solidFill'])
    if axis.spPr is None:
        axis.spPr = GraphicalProperties()
    axis.spPr.ln = ln


def _apply_color_to_prop(gp, color_spec, target_attr):
    """Apply a color spec to a graphical properties solidFill."""
    from openpyxl.drawing.fill import ColorChoice
    from openpyxl.drawing.colors import SchemeColor
    if isinstance(color_spec, dict) and 'scheme' in color_spec:
        sc = SchemeColor(val=color_spec['scheme'])
        if 'lumMod' in color_spec:
            sc.lumMod = color_spec['lumMod']
        if 'lumOff' in color_spec:
            sc.lumOff = color_spec['lumOff']
        setattr(gp, target_attr, ColorChoice(schemeClr=sc))
    elif isinstance(color_spec, str) and color_spec.startswith('#'):
        setattr(gp, target_attr, ColorChoice(srgbClr=color_spec[1:]))


def _apply_gridlines(axis, gl_spec):
    """Apply gridline formatting to a chart axis."""
    if not gl_spec:
        return
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    ln = LineProperties(w=gl_spec.get('width'), prstDash=gl_spec.get('prstDash'))
    if gl_spec.get('solidFill'):
        apply_chart_line_color(ln, gl_spec['solidFill'])
    sp = GraphicalProperties(ln=ln)
    axis.majorGridlines = ChartLines(spPr=sp)


def apply_chart_line_color(line_props, color_spec):
    """Apply color to a LineProperties solidFill."""
    from openpyxl.drawing.fill import ColorChoice
    if isinstance(color_spec, dict) and 'scheme' in color_spec:
        from openpyxl.drawing.colors import SchemeColor
        sc = SchemeColor(val=color_spec['scheme'])
        if 'lumMod' in color_spec:
            sc.lumMod = color_spec['lumMod']
        if 'lumOff' in color_spec:
            sc.lumOff = color_spec['lumOff']
        line_props.solidFill = ColorChoice(schemeClr=sc)
    elif isinstance(color_spec, str) and color_spec.startswith('#'):
        line_props.solidFill = ColorChoice(srgbClr=color_spec[1:])


def apply_chart_color(series, color_spec, target):
    from openpyxl.drawing.fill import ColorChoice
    if isinstance(color_spec, dict) and 'scheme' in color_spec:
        from openpyxl.drawing.colors import SchemeColor
        sc = SchemeColor(val=color_spec['scheme'])
        if 'lumMod' in color_spec:
            sc.lumMod = color_spec['lumMod']
        if 'lumOff' in color_spec:
            sc.lumOff = color_spec['lumOff']
        color = ColorChoice(schemeClr=sc)
    elif isinstance(color_spec, str) and color_spec.startswith('#'):
        color = ColorChoice(srgbClr=color_spec[1:])
    else:
        return
    if target == 'fill':
        series.graphicalProperties.solidFill = color
    elif target == 'line':
        from openpyxl.drawing.line import LineProperties
        if series.graphicalProperties.line is None:
            series.graphicalProperties.line = LineProperties()
        series.graphicalProperties.line.solidFill = color


def _copy_cell_style(src_cell, dst_cell):
    """Copy all style properties from src to dst cell. No serialization."""
    from copy import copy
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def _apply_ref_styles(wb, ref_path, spec, data):
    """Post-render pass: overwrite cell styles from reference xlsx directly.

    Uses role-based column mapping so that reference and target can have
    different numbers of asset columns. Columns are categorized as:
      - prefix (date cols before assets): mapped 1:1
      - asset (data columns): target asset cols get reference first-asset style
      - suffix (formula cols at end): mapped by offset from end
    """
    ref_wb = openpyxl.load_workbook(ref_path)

    for si, spec_sheet in enumerate(spec.get('sheets', [])):
        if si >= len(wb.sheetnames) or si >= len(ref_wb.sheetnames):
            continue
        tgt_ws = wb[wb.sheetnames[si]]
        ref_ws = ref_wb[ref_wb.sheetnames[si]]
        regions = spec_sheet.get('regions', [])
        if not regions:
            continue
        region = regions[0]

        n_header_rows = region.get('n_header_rows', 1)
        data_start = region.get('data_start_row', n_header_rows + 1)
        ref_start_col = region.get('start_col', 1)
        ref_end_col = region.get('end_col', 1)

        # Target sheet info
        tgt_data = data['sheets'][si] if si < len(data['sheets']) else None
        if not tgt_data:
            continue

        # Determine target's actual end column and data start row
        tgt_rows = tgt_data.get('rows', [])
        tgt_end_col = 0
        tgt_data_start = None
        for rd in tgt_rows:
            for c in rd.get('cells', []):
                tgt_end_col = max(tgt_end_col, c['col'])
            if rd['row_num'] >= data_start and tgt_data_start is None:
                tgt_data_start = rd['row_num']

        if tgt_end_col == 0:
            continue

        # Count suffix (formula) columns from target's first data row
        tgt_n_suffix = 0
        for rd in tgt_rows:
            if rd['row_num'] >= data_start:
                cells_desc = sorted(rd.get('cells', []), key=lambda x: x['col'], reverse=True)
                for c in cells_desc:
                    if c.get('is_formula'):
                        tgt_n_suffix += 1
                    else:
                        break
                break

        # Count suffix columns in reference using spec's last header row
        # (cells with non-numeric values after the dscore row)
        ref_n_suffix = 0
        hdr_rows = region.get('header_rows', [])
        if len(hdr_rows) >= 2:
            dscore_row = hdr_rows[-1]  # last header row has dscores
            # Find first non-numeric col after prefix (indicates formula area)
            sorted_cells = sorted(dscore_row, key=lambda c: c['col'])
            found_numeric = False
            for cell in sorted_cells:
                val = cell.get('value')
                if isinstance(val, (int, float)):
                    found_numeric = True
                elif found_numeric and val is not None:
                    # Non-numeric after numerics = start of formula labels
                    ref_n_suffix = ref_end_col - cell['col'] + 1
                    break

        # Determine asset range for reference and target
        # Prefix cols: cols before assets (typically 2: date, balance_date)
        # Detect from dscore row: first numeric value col = asset_start
        asset_start = ref_start_col + 2  # default
        if hdr_rows:
            for cell in sorted(hdr_rows[-1], key=lambda c: c['col']):
                if isinstance(cell.get('value'), (int, float)):
                    asset_start = cell['col']
                    break

        ref_asset_end = ref_end_col - ref_n_suffix
        tgt_asset_end = tgt_end_col - tgt_n_suffix

        # Build column mapping: target_col → reference_col
        col_map = {}
        # Prefix columns: direct 1:1
        for c in range(ref_start_col, asset_start):
            col_map[c] = c
        # Asset columns: map to reference first asset col (same style for all assets)
        first_asset_ref = asset_start
        for tc in range(asset_start, tgt_asset_end + 1):
            # Use corresponding ref asset col if in range, else first asset
            rc = tc if tc <= ref_asset_end else first_asset_ref
            col_map[tc] = rc
        # Suffix (formula) columns: map by offset from end
        for offset in range(max(tgt_n_suffix, ref_n_suffix)):
            tc = tgt_end_col - offset
            rc = ref_end_col - offset
            if tc >= asset_start and rc >= asset_start:
                col_map[tc] = rc

        # Apply header styles using column mapping
        for hi in range(n_header_rows):
            ref_r = region['start_row'] + hi
            tgt_r = region['start_row'] + hi
            if tgt_r > tgt_ws.max_row:
                break
            for tc in range(ref_start_col, tgt_end_col + 1):
                rc = col_map.get(tc)
                if rc is None or rc < ref_start_col or rc > ref_end_col:
                    continue
                ref_cell = ref_ws.cell(ref_r, rc)
                tgt_cell = tgt_ws.cell(tgt_r, tc)
                if tgt_cell.value is not None or ref_cell.value is not None:
                    _copy_cell_style(ref_cell, tgt_cell)

        # Apply data styles using column mapping (palette from ref first data row)
        if tgt_data_start:
            for rd in tgt_rows:
                r = rd['row_num']
                if r < tgt_data_start:
                    continue
                for cell_data in rd.get('cells', []):
                    tc = cell_data['col']
                    rc = col_map.get(tc)
                    if rc is None or rc < ref_start_col or rc > ref_end_col:
                        continue
                    ref_cell = ref_ws.cell(data_start, rc)
                    tgt_cell = tgt_ws.cell(r, tc)
                    _copy_cell_style(ref_cell, tgt_cell)

    ref_wb.close()


if __name__ == '__main__':
    main()