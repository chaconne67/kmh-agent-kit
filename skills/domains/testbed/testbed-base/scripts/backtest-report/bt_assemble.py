"""Assemble render input JSON from spec + params + CSV data."""
import argparse, json, pathlib, re, sys
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_dscore_map(asset_map):
    """asset_type -> dscore mapping from asset_map."""
    at_ds = {}
    for info in asset_map.values():
        at = info['asset_type']
        if at not in at_ds:
            at_ds[at] = info['dscore']
    return at_ds


def get_danger_map(asset_map):
    """asset_type -> is_danger mapping from asset_map."""
    at_d = {}
    for info in asset_map.values():
        at = info['asset_type']
        if at not in at_d:
            at_d[at] = info['is_danger']
    return at_d


def resolve_output_filename(params):
    """Determine output filename by scanning doc_dir for existing 14-* files."""
    doc_dir = pathlib.Path(params['doc_dir'])
    flavor = params['flavor']
    today = datetime.now().strftime('%Y%m%d')
    flavors = params.get('all_flavors', [flavor])

    flavor_num_map = {}
    for f in doc_dir.glob('*14-*백테스팅*'):
        m = re.search(r'14-(\d+)', f.name)
        if m:
            num = m.group(1)
            for fl in flavors:
                if fl in f.name:
                    flavor_num_map[fl] = num

    num = flavor_num_map.get(flavor, '1')
    folder_name = doc_dir.name
    return f'({folder_name}) 14-{num}.(자문일임)백테스팅결과분석자료_모멘텀에셋{flavor}_{today}.xlsx'


def _nan_to_none(val):
    """Convert NaN/None to None for JSON serialization."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    return val


# ---------------------------------------------------------------------------
# Generic simple sheet assembly (Sheets 1 & 2)
# ---------------------------------------------------------------------------

def assemble_simple_sheet(spec_sheet, params, csv_df, work_dir):
    """Generic assembly for simple sheets (Sheet 1, 2): header + data + chart.

    All headers, labels, and formula patterns come from spec_sheet.
    The CSV provides data values. The spec has a single region with:
    - header_rows: list of header row specs (first is labels, second may be data sample row)
    - data_sample: a sample data row for style extraction
    """
    regions = spec_sheet.get('regions', [])
    main_region = regions[0] if regions else {}

    rows = []

    # Determine header structure from spec
    spec_header_rows = main_region.get('header_rows', [])
    n_header_rows = main_region.get('n_header_rows', 1)

    # 1. Header row(s) — read labels from spec (only the label rows, not data sample rows)
    # The first header_row is always the label row
    if spec_header_rows:
        hdr_row = spec_header_rows[0]
        cells = []
        for cell_spec in hdr_row:
            cells.append({
                "col": cell_spec['col'],
                "value": cell_spec['value'],
                "style": cell_spec.get('style', {}),
            })
        if cells:
            rows.append({"row_num": 1, "cells": cells})

    # When spec has >1 header rows, check if header_row[1] is actually data mixed
    # with auxiliary cells (e.g. row 2 has dates at col 1 + aux label/formula at col 6-7).
    # In that case, real data starts at row 2 (not spec's data_start_row).
    # Auxiliary cells from header_row[1] are emitted separately.
    data_start = main_region.get('data_start_row', n_header_rows + 1)
    aux_cells_row2 = []

    if n_header_rows > 1 and len(spec_header_rows) > 1:
        csv_n_cols = len(csv_df.columns)
        hdr1 = spec_header_rows[1]
        # Identify aux cells: columns beyond CSV data range
        for cell_spec in hdr1:
            if cell_spec['col'] > csv_n_cols:
                aux_cells_row2.append({
                    "col": cell_spec['col'],
                    "value": cell_spec['value'],
                    "style": cell_spec.get('style', {}),
                    "is_formula": cell_spec.get('is_formula', False),
                })
        # Data starts right after the label row
        data_start = 2

    # Get styles from data_sample
    data_styles = {}
    for cell_spec in main_region.get('data_sample', []):
        data_styles[cell_spec['col']] = cell_spec.get('style', {})

    csv_cols = list(csv_df.columns)
    last_row = data_start + len(csv_df) - 1

    for idx, (_, row) in enumerate(csv_df.iterrows()):
        r = data_start + idx
        cells = []
        for col_idx, csv_col in enumerate(csv_cols):
            c = col_idx + 1
            val = _nan_to_none(row[csv_col])
            cell = {"col": c, "value": val, "style": data_styles.get(c, {})}
            # Detect date column from spec's number_format
            nf = data_styles.get(c, {}).get('number_format', '')
            if nf and ('yy' in str(nf).lower() or 'mm-dd' in str(nf).lower()):
                cell["value_type"] = "date"
            cells.append(cell)
        # Append aux cells from header_row[1] to the first data row
        if idx == 0 and aux_cells_row2:
            cells.extend(aux_cells_row2)
        rows.append({"row_num": r, "cells": cells})

    # 3. Chart — properties FROM SPEC, data references recomputed
    # Use series val_ref/cat_ref from spec to determine which columns are charted
    charts = []
    for chart_spec in spec_sheet.get('charts', []):
        chart = {**chart_spec}
        # Use spec's original anchor position (charts overlay on data, not below)
        chart["anchor"] = chart_spec.get('ref_anchor') or f"A{last_row + 2}"

        # Determine chart columns from spec series references
        spec_series = chart_spec.get('series', [])
        series_with_refs = [s for s in spec_series if 'val_ref' in s]

        if series_with_refs:
            # Use spec's series column references (adjust rows to actual data range)
            # The spec tells us WHICH columns to chart; we adjust row range to our data
            from openpyxl.utils import column_index_from_string
            val_cols = []
            for s in series_with_refs:
                vr = s['val_ref']
                val_cols.append(column_index_from_string(vr['min_col_letter']))
            cat_col = None
            first_cat = next((s.get('cat_ref') for s in series_with_refs if s.get('cat_ref')), None)
            if first_cat:
                cat_col = column_index_from_string(first_cat['min_col_letter'])

            # Build data_ref spanning the exact value columns
            chart["data_ref"] = {
                "min_col": min(val_cols),
                "max_col": max(val_cols),
                "min_row": 1,  # include header for titles
                "max_row": last_row,
            }
            chart["cat_ref"] = {
                "min_col": cat_col or 1,
                "min_row": data_start,
                "max_row": last_row,
            }
            # If series are non-contiguous (e.g. only col D out of B-G), use per-series refs
            if len(val_cols) == 1 or (max(val_cols) - min(val_cols) + 1) != len(val_cols):
                chart["per_series_refs"] = [
                    {"val_col": column_index_from_string(s['val_ref']['min_col_letter']),
                     "cat_col": column_index_from_string(s['cat_ref']['min_col_letter']) if s.get('cat_ref') else 1}
                    for s in series_with_refs
                ]
            chart["titles_from_data"] = True
        else:
            # Fallback: use all data columns
            chart["data_ref"] = {
                "min_col": main_region.get('start_col', 1) + 1,
                "max_col": main_region.get('end_col', 4),
                "min_row": main_region.get('start_row', 1),
                "max_row": last_row,
            }
            chart["cat_ref"] = {
                "min_col": main_region.get('start_col', 1),
                "min_row": data_start,
                "max_row": last_row,
            }
            chart["titles_from_data"] = True
        charts.append(chart)

    return {
        "name": spec_sheet['name'],
        "column_widths": spec_sheet.get('column_widths', {}),
        "row_heights": spec_sheet.get('row_heights', {}),
        "merged_cells": spec_sheet.get('merged_cells', []),
        "rows": rows,
        "charts": charts,
    }


# ---------------------------------------------------------------------------
# Sheet 1 & 2 wrappers
# ---------------------------------------------------------------------------

def assemble_sheet1(spec_sheet, params, asset_map, work_dir):
    """일별수익률추이 — delegates to generic assembly."""
    pk = params['pk']
    df = pd.read_csv(work_dir / f'{pk}_일별수익률.csv')
    return assemble_simple_sheet(spec_sheet, params, df, work_dir)


def assemble_sheet2(spec_sheet, params, asset_map, work_dir):
    """월별수익률 — delegates to generic assembly."""
    pk = params['pk']
    df = pd.read_csv(work_dir / f'{pk}_월별수익률.csv')
    return assemble_simple_sheet(spec_sheet, params, df, work_dir)


# ---------------------------------------------------------------------------
# Sheet 3: 자산종류별비중추이
# ---------------------------------------------------------------------------

def assemble_asset_sheet(spec_sheet, params, asset_map, work_dir, csv_name, add_cash=True):
    """Assembly for asset-based sheets (Sheet 3).

    Handles dynamic asset columns, dscore/danger computation, formula generation.
    Headers for fixed columns (합계, 위험자산비중, 위험도) come from spec.
    Asset column headers come from CSV column names (= DB asset_type).
    """
    pk = params['pk']
    df = pd.read_csv(work_dir / f'{pk}_{csv_name}.csv', index_col=0)

    dscore_map = get_dscore_map(asset_map)
    danger_map = get_danger_map(asset_map)

    # Cash config from params — not hardcoded
    cash_label = params.get('cash_label', '현금')
    cash_dscore = params.get('cash_dscore', 1)

    asset_cols = list(df.columns)
    if add_cash:
        if cash_label in asset_cols:
            # CSV already has cash column — don't recalculate
            pass
        else:
            df[cash_label] = 1.0 - df.sum(axis=1)
            asset_cols.append(cash_label)
    n_assets = len(asset_cols)

    # Compute dscores from asset_map
    dscores = []
    for at in asset_cols:
        if at == cash_label:
            dscores.append(cash_dscore)
        else:
            dscores.append(dscore_map.get(at, 0))

    # Compute danger flags from asset_map
    danger_flags = []
    for i, at in enumerate(asset_cols):
        if at == cash_label:
            danger_flags.append(False)
        elif at in danger_map:
            danger_flags.append(danger_map[at])
        else:
            danger_flags.append(dscores[i] >= 4)

    # Read spec structure
    regions = spec_sheet.get('regions', [])
    main_region = regions[0] if regions else {}
    spec_headers = main_region.get('header_rows', [[]])
    spec_data = main_region.get('data_sample', [])

    # Identify formula column labels from spec header row 0
    # In the spec, headers after the asset columns are formula columns (합계, 위험자산비중, 위험도)
    formula_col_labels = []
    for cell in spec_headers[0] if spec_headers else []:
        val = cell.get('value')
        if val is not None and isinstance(val, str) and cell['col'] > 2:
            # Check if this is a formula column label (not an asset type)
            # The spec has fixed labels like 합계, 위험자산비중, 위험도 at the end
            formula_col_labels.append({
                "label": val,
                "style": cell.get('style', {}),
                "original_col": cell['col'],
            })

    # Separate actual asset headers from formula headers
    # The spec header row has: col1=자산종류, cols 2-8=asset names, cols 9-11=formula labels
    # We need to identify which spec headers are formula columns vs asset columns
    # Formula columns in spec are identified by checking data_sample for formulas at those positions
    formula_spec_cols = set()
    for ds in spec_data:
        if ds.get('is_formula'):
            formula_spec_cols.add(ds['col'])

    # Also check the third header row (which is actually the first data row in spec)
    if len(spec_headers) > 2:
        for cell in spec_headers[2]:
            if cell.get('is_formula'):
                formula_spec_cols.add(cell['col'])

    # Re-identify formula column labels (only those whose columns have formulas)
    formula_col_labels = []
    for cell in spec_headers[0] if spec_headers else []:
        val = cell.get('value')
        if val is not None and isinstance(val, str) and cell['col'] in formula_spec_cols:
            formula_col_labels.append({
                "label": val,
                "style": cell.get('style', {}),
                "original_col": cell['col'],
            })

    # Column layout: A=날짜, B..B+n-1=assets, then formula columns
    sum_col = n_assets + 2
    danger_col = n_assets + 3
    risk_col = n_assets + 4
    cl = get_column_letter

    # Get styles from spec
    date_style = spec_data[0].get('style', {}) if spec_data else {}
    asset_data_style = spec_data[1].get('style', {}) if len(spec_data) > 1 else {}
    # Header style for asset columns: from spec header row (col 2+), not data row
    asset_hdr_style = asset_data_style  # fallback
    if spec_headers and spec_headers[0] and len(spec_headers[0]) > 1:
        asset_hdr_style = spec_headers[0][1].get('style', asset_data_style)
    # Formula styles: extract per-column styles from spec data_sample
    # Different formula columns may have different number_formats
    formula_style = {}  # generic (SUM/danger)
    risk_formula_style = {}  # SUMPRODUCT column
    for ds in spec_data:
        if ds.get('is_formula'):
            val = ds.get('value', '')
            if isinstance(val, str) and 'SUMPRODUCT' in val:
                risk_formula_style = ds.get('style', {})
            elif not formula_style:
                formula_style = ds.get('style', {})
    if not risk_formula_style:
        risk_formula_style = formula_style

    n_header_rows = main_region.get('n_header_rows', 2)
    data_start = main_region.get('data_start_row', n_header_rows + 1)

    # Detect if the last header_row is actually a data row (contains dates at col 1).
    # If so, adjust data_start to include it in the data range.
    if n_header_rows > 2 and len(spec_headers) >= n_header_rows:
        last_hdr = spec_headers[n_header_rows - 1]
        if last_hdr:
            first_val = last_hdr[0].get('value', '')
            if isinstance(first_val, str) and re.match(r'\d{4}-\d{2}-\d{2}', first_val):
                data_start = n_header_rows  # row 3 instead of 4

    rows = []

    # Row 1: header labels from spec + asset names from CSV
    hdr1_cells = []
    # First cell: from spec (e.g. '자산종류')
    if spec_headers and spec_headers[0]:
        hdr1_cells.append({
            "col": 1,
            "value": spec_headers[0][0].get('value'),
            "style": spec_headers[0][0].get('style', {}),
        })
    for i, at in enumerate(asset_cols):
        hdr1_cells.append({"col": i + 2, "value": at, "style": asset_hdr_style})
    # Formula column headers from spec
    for idx, fc in enumerate(formula_col_labels):
        target_col = sum_col + idx
        hdr1_cells.append({"col": target_col, "value": fc['label'], "style": fc['style']})
    rows.append({"row_num": 1, "cells": hdr1_cells})

    # Row 2: dscore values + date label from spec
    hdr2_cells = []
    if len(spec_headers) > 1 and spec_headers[1]:
        hdr2_cells.append({
            "col": 1,
            "value": spec_headers[1][0].get('value'),
            "style": spec_headers[1][0].get('style', {}),
        })
    dscore_style = spec_headers[1][1].get('style', {}) if len(spec_headers) > 1 and len(spec_headers[1]) > 1 else {}
    for i, ds in enumerate(dscores):
        hdr2_cells.append({"col": i + 2, "value": ds, "style": dscore_style})
    rows.append({"row_num": 2, "cells": hdr2_cells})

    # Merged cells: formula columns span rows 1-2
    merged = []
    for idx in range(len(formula_col_labels)):
        target_col = sum_col + idx
        merged.append(f"{cl(target_col)}1:{cl(target_col)}2")

    # Data rows
    last_row = data_start + len(df) - 1
    for idx, (date_str, row) in enumerate(df.iterrows()):
        r = data_start + idx
        cells = [{"col": 1, "value": str(date_str), "value_type": "date", "style": date_style}]
        for i, at in enumerate(asset_cols):
            cells.append({"col": i + 2, "value": float(row[at]), "style": asset_data_style})
        # Sum formula
        cells.append({
            "col": sum_col,
            "value": f"=SUM({cl(2)}{r}:{cl(n_assets + 1)}{r})",
            "is_formula": True,
            "style": formula_style,
        })
        # Danger sum
        danger_parts = [f"{cl(i + 2)}{r}" for i in range(n_assets) if danger_flags[i]]
        if danger_parts:
            cells.append({
                "col": danger_col,
                "value": f'={"+".join(danger_parts)}',
                "is_formula": True,
                "style": formula_style,
            })
        else:
            cells.append({"col": danger_col, "value": 0, "style": formula_style})
        # Risk SUMPRODUCT
        cells.append({
            "col": risk_col,
            "value": f"=SUMPRODUCT($B$2:${cl(n_assets + 1)}$2,B{r}:{cl(n_assets + 1)}{r})",
            "is_formula": True,
            "style": risk_formula_style,
        })
        rows.append({"row_num": r, "cells": cells})

    # Aux data from spec — check for extra cells in header_rows or data_sample
    # The spec has cells at columns L-O with things like '자산별 최고치' and MAX formula
    aux_cells_by_row = defaultdict(list)
    # Look at the third header row (first data row in spec) for aux cells
    if len(spec_headers) > 2:
        for cell in spec_headers[2]:
            if cell['col'] > risk_col and cell.get('value') is not None:
                aux_cells_by_row[data_start].append({
                    "col": cell['col'],
                    "value": cell['value'],
                    "style": cell.get('style', {}),
                    "is_formula": cell.get('is_formula', False),
                })
    # Check data_sample for aux cells beyond formula columns
    for cell in spec_data:
        if cell['col'] > risk_col and cell.get('value') is not None:
            aux_cells_by_row[data_start + 1].append({
                "col": cell['col'],
                "value": cell['value'],
                "style": cell.get('style', {}),
                "is_formula": cell.get('is_formula', False),
            })

    for r, cells in sorted(aux_cells_by_row.items()):
        rows.append({"row_num": r, "cells": cells})

    return {
        "name": spec_sheet['name'],
        "column_widths": spec_sheet.get('column_widths', {}),
        "row_heights": spec_sheet.get('row_heights', {}),
        "merged_cells": merged,
        "rows": rows,
        "charts": [],
    }


def assemble_sheet3(spec_sheet, params, asset_map, work_dir):
    return assemble_asset_sheet(spec_sheet, params, asset_map, work_dir,
                                csv_name='자산종류별비중', add_cash=True)


# ---------------------------------------------------------------------------
# Sheet 4: 종목별투자비중추이
# ---------------------------------------------------------------------------

def assemble_sheet4(spec_sheet, params, asset_map, work_dir):
    """종목별투자비중추이: ticker data + AreaChart percentStacked.

    Headers = ticker codes from CSV columns (not hardcoded).
    Styles from spec's data_sample.
    Chart type/grouping/style from spec, series colors cycle through spec's series list.
    """
    pk = params['pk']
    df = pd.read_csv(work_dir / f'{pk}_종목별비중.csv', index_col=0)
    ticker_cols = list(df.columns)
    n_tickers = len(ticker_cols)

    regions = spec_sheet.get('regions', [])
    main_region = regions[0] if regions else {}
    spec_headers = main_region.get('header_rows', [[]])
    spec_data = main_region.get('data_sample', [])

    # Styles from spec
    hdr_style = spec_headers[0][0].get('style', {}) if spec_headers and spec_headers[0] else {}
    date_style = spec_data[0].get('style', {}) if spec_data else {}
    data_style = spec_data[1].get('style', {}) if len(spec_data) > 1 else {}

    rows = []
    # Header row: col1 from spec, then ticker codes from CSV
    hdr_cells = [{"col": 1, "value": spec_headers[0][0]['value'], "style": hdr_style}]
    for i, tc in enumerate(ticker_cols):
        hdr_cells.append({"col": i + 2, "value": tc, "style": hdr_style})
    rows.append({"row_num": 1, "cells": hdr_cells})

    # Data rows
    data_start = main_region.get('data_start_row', 2)
    last_row = data_start + len(df) - 1
    for idx, (date_str, row) in enumerate(df.iterrows()):
        r = data_start + idx
        cells = [{"col": 1, "value": str(date_str), "value_type": "date", "style": date_style}]
        for i, tc in enumerate(ticker_cols):
            cells.append({"col": i + 2, "value": float(row[tc]), "style": data_style})
        rows.append({"row_num": r, "cells": cells})

    # Chart from spec, with updated data references
    charts = []
    for chart_spec in spec_sheet.get('charts', []):
        chart = {**chart_spec}
        chart["anchor"] = chart_spec.get('ref_anchor') or f"A{last_row + 2}"
        chart["data_ref"] = {
            "min_col": 2,
            "max_col": n_tickers + 1,
            "min_row": 1,
            "max_row": last_row,
        }
        chart["cat_ref"] = {"min_col": 1, "min_row": 2, "max_row": last_row}
        chart["titles_from_data"] = True
        # If actual series count > spec series count, cycle spec colors
        spec_series = chart_spec.get('series', [])
        if spec_series and n_tickers > len(spec_series):
            extended = []
            for i in range(n_tickers):
                extended.append(spec_series[i % len(spec_series)])
            chart["series"] = extended
        charts.append(chart)

    return {
        "name": spec_sheet['name'],
        "column_widths": spec_sheet.get('column_widths', {}),
        "row_heights": spec_sheet.get('row_heights', {}),
        "merged_cells": [],
        "rows": rows,
        "charts": charts,
    }


# ---------------------------------------------------------------------------
# Sheet 5: 리밸런싱발생내역
# ---------------------------------------------------------------------------

def assemble_sheet5(spec_sheet, params, asset_map, work_dir):
    """리밸런싱발생내역: meta + 2-row header + data + formulas.

    The spec has a single region with 5 header rows:
      - Row 0: meta labels (포트폴리오유형, 모아포트폴리오리밸런싱, ...)
      - Row 1: meta values (적극투자형, 중간맛, 0.75, 1, formulas)
      - Row 2: empty spacer
      - Row 3: data table header row 1 (자산종류, asset names, formula col labels)
      - Row 4: data table header row 2 (리밸런싱날짜, 단기만료일도래, dscores, sub-labels)
    Then data_sample is the first data row.

    Params values are used AS-IS via direct reference — never re-derived.
    """
    pk = params['pk']
    cash_label = params.get('cash_label', '현금')
    cash_dscore = params.get('cash_dscore', 1)

    df = pd.read_csv(work_dir / f'{pk}_리밸런싱내역.csv', index_col=0)

    dscore_map = get_dscore_map(asset_map)
    danger_map = get_danger_map(asset_map)

    asset_cols = list(df.columns)
    if cash_label in asset_cols:
        # CSV already has cash column — don't recalculate
        pass
    else:
        df[cash_label] = 1.0 - df.sum(axis=1)
        asset_cols.append(cash_label)
    n_assets = len(asset_cols)

    dscores = [cash_dscore if at == cash_label else dscore_map.get(at, 0) for at in asset_cols]
    danger_flags = [
        False if at == cash_label
        else (danger_map.get(at, dscores[i] >= 4))
        for i, at in enumerate(asset_cols)
    ]

    # Read spec structure — single region with 5 header rows
    regions = spec_sheet.get('regions', [])
    main_region = regions[0] if regions else {}
    spec_headers = main_region.get('header_rows', [])
    spec_data = main_region.get('data_sample', [])
    spec_formulas = main_region.get('formulas', [])

    cl = get_column_letter

    # The spec has a fixed structure:
    # Meta is in rows 1-2 (header_rows[0] and [1])
    # Data table headers in rows 4-5 (header_rows[3] and [4])
    # Data starts at row 6 (data_start_row)

    meta_labels_row = spec_headers[0] if len(spec_headers) > 0 else []
    meta_values_row = spec_headers[1] if len(spec_headers) > 1 else []
    data_hdr_row1 = spec_headers[3] if len(spec_headers) > 3 else []
    data_hdr_row2 = spec_headers[4] if len(spec_headers) > 4 else []

    # Derive asset_start_col from spec: count non-asset header cells in data_hdr_row2
    # (cells before the first numeric dscore value)
    non_asset_count = 0
    for cell in sorted(data_hdr_row2, key=lambda c: c['col']):
        val = cell.get('value')
        if isinstance(val, (int, float)):
            break
        non_asset_count += 1
    asset_start_col = non_asset_count + 1

    # Formula columns follow assets
    danger_sum_col = asset_start_col + n_assets
    risk_col = danger_sum_col + 1
    reason_col = risk_col + 1
    eval_col1 = reason_col + 1
    eval_col2 = eval_col1 + 1
    disqual_col = eval_col2 + 1

    rows = []

    # --- Row 1: Meta labels FROM SPEC ---
    meta_label_cells = []
    for cell_spec in meta_labels_row:
        val = cell_spec.get('value')
        if val is not None:
            meta_label_cells.append({
                "col": cell_spec['col'],
                "value": val,
                "style": cell_spec.get('style', {}),
            })
    rows.append({"row_num": 1, "cells": meta_label_cells})

    # --- Row 2: Meta values FROM PARAMS (direct reference, no re-derivation) ---
    # Map spec header labels to params keys
    meta_param_map = params.get('meta_param_map', {})
    meta_value_cells = []
    meta_style = meta_values_row[0].get('style', {}) if meta_values_row else {}

    if meta_param_map:
        # Explicit label → param key mapping from params.json
        for cell_spec in meta_values_row:
            val = cell_spec.get('value')
            if val is None:
                continue

            # Find the corresponding label from row 1
            label = None
            for lbl_cell in meta_labels_row:
                if lbl_cell['col'] == cell_spec['col'] and lbl_cell.get('value') is not None:
                    label = lbl_cell['value']
                    break

            if cell_spec.get('is_formula'):
                meta_value_cells.append({
                    "col": cell_spec['col'],
                    "value": cell_spec['value'],
                    "is_formula": True,
                    "style": cell_spec.get('style', meta_style),
                })
            elif label and meta_param_map.get(label) in params:
                meta_value_cells.append({
                    "col": cell_spec['col'],
                    "value": params[meta_param_map[label]],
                    "style": cell_spec.get('style', meta_style),
                })
            else:
                meta_value_cells.append({
                    "col": cell_spec['col'],
                    "value": val,
                    "style": cell_spec.get('style', meta_style),
                })
    else:
        raise ValueError(
            "Missing 'meta_param_map' in bt_params.json. "
            "Please add a 'meta_param_map' dict that maps Sheet 5 meta label strings "
            "(from the spec's header_rows[0]) to param keys in bt_params.json. "
            "Example: {\"포트폴리오유형\": \"portfolio_type\", \"모아포트폴리오리밸런싱\": \"flavor\", ...}"
        )

    rows.append({"row_num": 2, "cells": meta_value_cells})

    # --- Row 3: empty spacer (skip) ---

    # --- Data table header row 1 ---
    # Derive header start: meta occupies 2 rows + 1 spacer row, then data headers
    # Use spec's data_start_row and n_header_rows to compute:
    # data headers occupy (n_header_rows - meta rows) rows before data_start_row
    n_meta_rows = 2  # header_rows[0] and [1] are meta
    n_data_hdr_rows = main_region.get('n_header_rows', 5) - n_meta_rows - 1  # minus spacer
    data_hdr_start = main_region.get('data_start_row', 6) - n_data_hdr_rows
    hdr1_style = data_hdr_row1[0].get('style', {}) if data_hdr_row1 else {}
    hdr1_cells = []

    # Fixed header cells from spec (cols before asset_start_col)
    for cell_spec in data_hdr_row1:
        if cell_spec['col'] < asset_start_col and cell_spec.get('value') is not None:
            hdr1_cells.append({
                "col": cell_spec['col'],
                "value": cell_spec['value'],
                "style": cell_spec.get('style', hdr1_style),
            })

    # Asset column headers from CSV
    for i, at in enumerate(asset_cols):
        hdr1_cells.append({
            "col": asset_start_col + i,
            "value": at,
            "style": hdr1_style,
        })

    # Formula column headers from spec — identify where formula zone starts
    # by finding the first formula column (after date cols) in data_sample.
    # Everything in the header at or after that column is a formula/extra label.
    formula_zone_start = None
    for ds_cell in sorted(spec_data, key=lambda c: c['col']):
        if ds_cell.get('is_formula') and ds_cell['col'] > asset_start_col:
            formula_zone_start = ds_cell['col']
            break

    formula_labels_from_spec = []
    if formula_zone_start:
        for cell_spec in sorted(data_hdr_row1, key=lambda c: c['col']):
            val = cell_spec.get('value')
            if val is None or not isinstance(val, str):
                continue
            if cell_spec['col'] >= formula_zone_start:
                formula_labels_from_spec.append({
                    "label": val,
                    "style": cell_spec.get('style', hdr1_style),
                    "original_col": cell_spec['col'],
                })

    # eval_col2 is omitted: "평" header spans eval_col1:eval_col2 via merge
    target_cols = [danger_sum_col, risk_col, reason_col, eval_col1, disqual_col]
    for i, fc in enumerate(formula_labels_from_spec):
        if i < len(target_cols):
            hdr1_cells.append({
                "col": target_cols[i],
                "value": fc['label'],
                "style": fc.get('style', hdr1_style),
            })
    rows.append({"row_num": data_hdr_start, "cells": hdr1_cells})

    # --- Row 5: Sub-headers (date labels + dscores) ---
    hdr2_style = data_hdr_row2[0].get('style', {}) if data_hdr_row2 else {}
    hdr2_cells = []
    for cell_spec in data_hdr_row2:
        val = cell_spec.get('value')
        if val is not None and cell_spec['col'] < asset_start_col:
            hdr2_cells.append({
                "col": cell_spec['col'],
                "value": val,
                "style": cell_spec.get('style', hdr2_style),
            })
    for i, ds in enumerate(dscores):
        hdr2_cells.append({
            "col": asset_start_col + i,
            "value": ds,
            "style": hdr2_style,
        })

    # Sub-labels for eval columns from spec — collect string labels from
    # data_hdr_row2 that are in the formula zone (at or beyond formula_zone_start)
    eval_sub_labels = []
    for cell_spec in data_hdr_row2:
        val = cell_spec.get('value')
        if val is not None and isinstance(val, str) and formula_zone_start and cell_spec['col'] >= formula_zone_start:
            eval_sub_labels.append({
                "value": val,
                "style": cell_spec.get('style', hdr2_style),
            })
    eval_target_cols = [eval_col1, eval_col2]
    for i, sub in enumerate(eval_sub_labels):
        if i < len(eval_target_cols):
            hdr2_cells.append({
                "col": eval_target_cols[i],
                "value": sub['value'],
                "style": sub['style'],
            })

    rows.append({"row_num": data_hdr_start + 1, "cells": hdr2_cells})

    # --- Merged cells (all computed dynamically, never copied from spec) ---
    merged = []
    # Fixed columns A:B merge across data header rows (자산종류 spans 2 cols)
    if asset_start_col > 2:
        merged.append(f"A{data_hdr_start}:{cl(asset_start_col - 1)}{data_hdr_start}")
    # Formula columns that span both header rows (danger_sum, risk, reason, disqual)
    for tc in [danger_sum_col, risk_col, reason_col, disqual_col]:
        merged.append(f"{cl(tc)}{data_hdr_start}:{cl(tc)}{data_hdr_start + 1}")
    # Eval columns span header row 1 only (row 4), with sub-labels in row 2
    merged.append(f"{cl(eval_col1)}{data_hdr_start}:{cl(eval_col2)}{data_hdr_start}")

    # --- Data rows ---
    data_start = main_region.get('data_start_row', 6)
    date_style = spec_data[0].get('style', {}) if spec_data else {}
    pct_style = spec_data[2].get('style', {}) if len(spec_data) > 2 else {}
    # Extract per-formula-type styles from spec data_sample
    # Skip date formulas (col <= 2) when identifying formula styles
    formula_style_ds = {}
    risk_formula_style_ds = {}
    for ds_cell in spec_data:
        if ds_cell.get('is_formula') and ds_cell.get('col', 0) > 2:
            val = ds_cell.get('value', '')
            if isinstance(val, str) and 'SUMPRODUCT' in val:
                risk_formula_style_ds = ds_cell.get('style', {})
            elif not formula_style_ds:
                formula_style_ds = ds_cell.get('style', {})
    if not risk_formula_style_ds:
        risk_formula_style_ds = formula_style_ds

    dscore_range = f"${cl(asset_start_col)}${data_hdr_start + 1}:${cl(asset_start_col + n_assets - 1)}${data_hdr_start + 1}"

    # Disqual style: use the last cell in spec data_sample (the disqual column)
    disqual_style = formula_style_ds
    if spec_data:
        last_spec_cell = max(spec_data, key=lambda c: c['col'])
        disqual_style = last_spec_cell.get('style', formula_style_ds)

    # Reason value: read from spec's data sample (never hardcode)
    # Skip date-like strings and formula cells; look for short text labels
    reason_val = None
    for ds_cell in spec_data:
        if ds_cell.get('col') and not ds_cell.get('is_formula'):
            val = ds_cell.get('value')
            if isinstance(val, str) and val not in asset_cols:
                # Skip date-like values
                if re.match(r'\d{4}-\d{2}-\d{2}', val):
                    continue
                reason_val = val
                break
    if reason_val is None:
        reason_val = ""  # Safe fallback

    # Find meta column positions for eval formulas by matching spec meta values
    # against params values. The meta_values_row contains the actual values
    # (risk_limit, min_risk) and a formula for max_risk.
    rl_col_letter = None
    mr_col_letter = None
    mx_col_letter = None

    if meta_param_map:
        # Use meta_param_map to find which labels correspond to which params
        for cell_spec in meta_labels_row:
            label = cell_spec.get('value')
            if label is None:
                continue
            param_key = meta_param_map.get(label)
            if param_key == 'risk_limit':
                rl_col_letter = cl(cell_spec['col'])
            elif param_key == 'min_risk':
                mr_col_letter = cl(cell_spec['col'])
            elif param_key == 'max_risk':
                mx_col_letter = cl(cell_spec['col'])
    else:
        # Positional approach: non-string, non-formula cells in meta_values_row
        # correspond to [risk_limit, min_risk] in order (strings are portfolio_type/flavor)
        numeric_meta_cols = []
        for cell_spec in meta_values_row:
            val = cell_spec.get('value')
            if val is None or cell_spec.get('is_formula'):
                continue
            if not isinstance(val, str):
                numeric_meta_cols.append(cell_spec['col'])
        if len(numeric_meta_cols) >= 1:
            rl_col_letter = cl(numeric_meta_cols[0])
        if len(numeric_meta_cols) >= 2:
            mr_col_letter = cl(numeric_meta_cols[1])

    # max_risk is typically a formula cell in meta row 2 (fallback only)
    if mx_col_letter is None:
        for cell_spec in meta_values_row:
            if cell_spec.get('is_formula') and cell_spec.get('value'):
                mx_col_letter = cl(cell_spec['col'])
                break

    if not all([rl_col_letter, mr_col_letter, mx_col_letter]):
        raise ValueError(
            f"Could not determine meta column positions for eval formulas. "
            f"Found: risk_limit={rl_col_letter}, min_risk={mr_col_letter}, max_risk={mx_col_letter}. "
            f"Provide 'meta_param_map' in params.json to map meta labels to param keys."
        )

    for idx, (date_str, row) in enumerate(df.iterrows()):
        r = data_start + idx
        cells = [
            {"col": 1, "value": str(date_str), "value_type": "date", "style": date_style},
            {"col": 2, "value": f"=A{r}+1", "is_formula": True, "style": date_style},
        ]
        for i, at in enumerate(asset_cols):
            cells.append({
                "col": asset_start_col + i,
                "value": float(row[at]),
                "style": pct_style,
            })
        # Danger sum
        danger_parts = [f"{cl(asset_start_col + i)}{r}" for i in range(n_assets) if danger_flags[i]]
        cells.append({
            "col": danger_sum_col,
            "value": f'={"+".join(danger_parts)}' if danger_parts else "0",
            "is_formula": bool(danger_parts),
            "style": formula_style_ds,
        })
        # Risk SUMPRODUCT
        cells.append({
            "col": risk_col,
            "value": f"=SUMPRODUCT({dscore_range},{cl(asset_start_col)}{r}:{cl(asset_start_col + n_assets - 1)}{r})",
            "is_formula": True,
            "style": risk_formula_style_ds,
        })
        # Reason — from spec, never hardcoded
        cells.append({"col": reason_col, "value": reason_val, "style": pct_style})
        # Eval formulas
        dc = cl(danger_sum_col)
        rk = cl(risk_col)
        cells.append({
            "col": eval_col1,
            "value": f'=IF({dc}{r}="","",IF(${rl_col_letter}$2>={dc}{r},"O","X"))',
            "is_formula": True,
            "style": formula_style_ds,
        })
        cells.append({
            "col": eval_col2,
            "value": f'=IF({rk}{r}="","",IF(AND(${mr_col_letter}$2<={rk}{r},{rk}{r}<=${mx_col_letter}$2),"O","X"))',
            "is_formula": True,
            "style": formula_style_ds,
        })
        # Disqualification formula: check if both eval columns are "O"
        e1 = cl(eval_col1)
        e2 = cl(eval_col2)
        cells.append({
            "col": disqual_col,
            "value": f'=IF(AND({e1}{r}="O",{e2}{r}="O"),"적격","부적격")',
            "is_formula": True,
            "style": disqual_style,
        })
        rows.append({"row_num": r, "cells": cells})

    return {
        "name": spec_sheet['name'],
        "column_widths": spec_sheet.get('column_widths', {}),
        "row_heights": spec_sheet.get('row_heights', {}),
        "merged_cells": merged,
        "rows": rows,
        "charts": [],
    }


# ---------------------------------------------------------------------------
# Main assemble
# ---------------------------------------------------------------------------

def assemble(spec, params, asset_map, work_dir):
    output_filename = resolve_output_filename(params)

    assemblers = [
        assemble_sheet1,
        assemble_sheet2,
        assemble_sheet3,
        assemble_sheet4,
        assemble_sheet5,
    ]

    sheets = []
    for i, spec_sheet in enumerate(spec['sheets']):
        if i < len(assemblers):
            sheet_data = assemblers[i](spec_sheet, params, asset_map, work_dir)
            sheets.append(sheet_data)

    return {
        "output_filename": output_filename,
        "sheets": sheets,
    }


def main():
    parser = argparse.ArgumentParser(description='Assemble bt_render_input.json')
    parser.add_argument('--spec', required=True, help='bt_spec.json path')
    parser.add_argument('--params', required=True, help='bt_params.json path')
    parser.add_argument('--asset-map', required=True, help='bt_asset_map.json path')
    parser.add_argument('--work-dir', required=True, help='Working directory with CSV files')
    parser.add_argument('--output', required=True, help='Output bt_render_input.json path')
    args = parser.parse_args()

    spec = json.loads(pathlib.Path(args.spec).read_text(encoding='utf-8'))
    params = json.loads(pathlib.Path(args.params).read_text(encoding='utf-8'))
    asset_map = json.loads(pathlib.Path(args.asset_map).read_text(encoding='utf-8'))
    work_dir = pathlib.Path(args.work_dir)

    render_input = assemble(spec, params, asset_map, work_dir)

    out_path = pathlib.Path(args.output)
    out_path.write_text(json.dumps(render_input, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Render input assembled: {len(render_input["sheets"])} sheets -> {out_path}')


if __name__ == '__main__':
    main()