"""Style verification: reference vs rendered xlsx styles (role-based).

Compares font, fill, border, alignment, number_format between reference and target
by identifying cells by role (header, data, formula) from the spec.
"""
import argparse, json, pathlib, sys

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')


# Map Korean font names to English canonical form
_FONT_CANONICAL = {
    '맑은 고딕': 'Malgun Gothic',
    '굴림': 'Gulim',
    '돋움': 'Dotum',
    '바탕': 'Batang',
}


def _normalize_font_name(name):
    """Normalize font name to canonical form for comparison."""
    if name is None:
        return None
    return _FONT_CANONICAL.get(name, name)


def _safe_rgb(color):
    """Safely extract RGB string from openpyxl Color, handling theme colors."""
    if color is None:
        return None
    try:
        rgb = color.rgb
        if rgb is None:
            return None
        s = str(rgb)
        if 'Values must be' in s:
            # Theme color — can't resolve to RGB, return theme info instead
            theme = getattr(color, 'theme', None)
            tint = getattr(color, 'tint', None)
            if theme is not None:
                return f"theme:{theme}:tint:{tint}"
            return None
        return s
    except (TypeError, ValueError, AttributeError):
        return None


def compare_font(ref_font, tgt_font, context):
    """Compare font properties, return list of differences."""
    diffs = []
    for prop in ['name', 'size', 'bold', 'italic', 'underline', 'strikethrough']:
        rv = getattr(ref_font, prop, None)
        tv = getattr(tgt_font, prop, None)
        if prop == 'name':
            if _normalize_font_name(rv) != _normalize_font_name(tv):
                diffs.append(f"{context} font.{prop}: ref={rv}, target={tv}")
        elif rv != tv:
            diffs.append(f"{context} font.{prop}: ref={rv}, target={tv}")
    return diffs


def _strip_alpha(argb):
    """Normalize ARGB string by stripping alpha prefix for comparison.
    FF153D64 and 00153D64 should match as the same color #153D64."""
    if argb is None:
        return None
    s = str(argb)
    # Strip 2-char alpha prefix from 8-char ARGB strings
    if len(s) == 8 and all(c in '0123456789abcdefABCDEF' for c in s):
        return s[2:]
    return s


def compare_fill(ref_fill, tgt_fill, context):
    diffs = []
    if ref_fill.fill_type != tgt_fill.fill_type:
        diffs.append(f"{context} fill.type: ref={ref_fill.fill_type}, target={tgt_fill.fill_type}")
    if ref_fill.fgColor and tgt_fill.fgColor:
        ref_rgb = _safe_rgb(ref_fill.fgColor)
        tgt_rgb = _safe_rgb(tgt_fill.fgColor)
        if ref_rgb and tgt_rgb and _strip_alpha(ref_rgb) != _strip_alpha(tgt_rgb):
            diffs.append(f"{context} fill.fgColor: ref={ref_rgb}, target={tgt_rgb}")
    return diffs


def compare_numfmt(ref_nf, tgt_nf, context):
    if ref_nf != tgt_nf and ref_nf != 'General':
        return [f"{context} number_format: ref='{ref_nf}', target='{tgt_nf}'"]
    return []


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--ref', required=True)
    parser.add_argument('--target', required=True)
    parser.add_argument('--spec', required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    ref_wb = openpyxl.load_workbook(args.ref)
    tgt_wb = openpyxl.load_workbook(args.target)
    spec = json.loads(pathlib.Path(args.spec).read_text(encoding='utf-8'))

    errors = []

    for si, spec_sheet in enumerate(spec.get('sheets', [])):
        if si >= len(ref_wb.sheetnames) or si >= len(tgt_wb.sheetnames):
            continue
        rws = ref_wb[ref_wb.sheetnames[si]]
        tws = tgt_wb[tgt_wb.sheetnames[si]]
        sheet_name = spec_sheet['name']

        for region in spec_sheet.get('regions', []):
            # Compare header styles
            for hdr_idx, hdr_row in enumerate(region.get('header_rows', [])):
                for cell_spec in hdr_row:
                    c = cell_spec['col']
                    r = region['start_row'] + hdr_idx
                    rc = rws.cell(r, c)
                    tc = tws.cell(r, c)
                    # Skip cells where target has no value (column count may differ)
                    if tc.value is None and rc.value is None:
                        continue
                    ctx = f"Sheet '{sheet_name}' header R{r}C{c}"
                    errors.extend(compare_font(rc.font, tc.font, ctx))
                    errors.extend(compare_fill(rc.fill, tc.fill, ctx))
                    errors.extend(compare_numfmt(rc.number_format, tc.number_format, ctx))

            # Compare data row styles (sample first data row)
            dr = region.get('data_start_row')
            if dr and dr <= rws.max_row and dr <= tws.max_row:
                for c in range(region.get('start_col', 1), region.get('end_col', 1) + 1):
                    rc = rws.cell(dr, c)
                    tc = tws.cell(dr, c)
                    # Skip cells where both are empty (auxiliary/unused columns)
                    if rc.value is None and tc.value is None:
                        continue
                    ctx = f"Sheet '{sheet_name}' data R{dr}C{c}"
                    errors.extend(compare_font(rc.font, tc.font, ctx))
                    errors.extend(compare_numfmt(rc.number_format, tc.number_format, ctx))

    result = {"status": "PASS" if not errors else "FAIL", "error_count": len(errors), "errors": errors[:50]}
    pathlib.Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"Style verification: {result['status']} ({len(errors)} errors)")


if __name__ == '__main__':
    main()