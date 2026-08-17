"""Structure verification: reference vs rendered xlsx layout."""
import argparse, json, pathlib, sys

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--ref', required=True)
    parser.add_argument('--target', required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    ref_wb = openpyxl.load_workbook(args.ref)
    tgt_wb = openpyxl.load_workbook(args.target)

    errors = []

    # Sheet count
    if len(ref_wb.sheetnames) != len(tgt_wb.sheetnames):
        errors.append(f"Sheet count: ref={len(ref_wb.sheetnames)}, target={len(tgt_wb.sheetnames)}")

    # Sheet names
    for i, (rn, tn) in enumerate(zip(ref_wb.sheetnames, tgt_wb.sheetnames)):
        if rn != tn:
            errors.append(f"Sheet {i} name: ref='{rn}', target='{tn}'")

    # Per-sheet: chart count and type
    for i, (rn, tn) in enumerate(zip(ref_wb.sheetnames, tgt_wb.sheetnames)):
        rws = ref_wb[rn]
        tws = tgt_wb[tn]
        if len(rws._charts) != len(tws._charts):
            errors.append(f"Sheet '{rn}' chart count: ref={len(rws._charts)}, target={len(tws._charts)}")
        for j, (rc, tc) in enumerate(zip(rws._charts, tws._charts)):
            if type(rc).__name__ != type(tc).__name__:
                errors.append(f"Sheet '{rn}' chart {j} type: ref={type(rc).__name__}, target={type(tc).__name__}")

    result = {"status": "PASS" if not errors else "FAIL", "errors": errors}
    pathlib.Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"Structure verification: {result['status']} ({len(errors)} errors)")


if __name__ == '__main__':
    main()