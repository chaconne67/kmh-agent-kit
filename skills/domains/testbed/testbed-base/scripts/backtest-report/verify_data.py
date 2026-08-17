"""Data verification: source data vs rendered xlsx cell values.

Verifies ALL sheets by comparing source CSV data against rendered xlsx cells.
Column names are read dynamically from CSV headers — no hardcoded column names.
"""
import argparse, json, pathlib, sys
from datetime import datetime

import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

TOLERANCE = 0.001


def verify_csv_sheet(ws, df, header_row, data_start_row, errors, sheet_label):
    """Generic: verify CSV data columns against xlsx cells."""
    csv_cols = list(df.columns)
    for col_idx, col_name in enumerate(csv_cols):
        c = col_idx + 1
        for row_idx, (_, row) in enumerate(df.iterrows()):
            r = data_start_row + row_idx
            cell_val = ws.cell(r, c).value
            expected = row[col_name]
            # Skip formula cells (they reference other cells)
            if isinstance(cell_val, str) and cell_val.startswith('='):
                continue
            # Compare numeric values with tolerance
            try:
                if cell_val is not None and abs(float(cell_val) - float(expected)) > TOLERANCE:
                    errors.append(f"{sheet_label} row{r} col{c}({col_name}): expected {expected}, got {cell_val}")
            except (ValueError, TypeError):
                # Date or string comparison
                if str(cell_val)[:10] != str(expected)[:10]:
                    errors.append(f"{sheet_label} row{r} col{c}({col_name}): expected {expected}, got {cell_val}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--params', required=True)
    parser.add_argument('--asset-map', required=True)
    parser.add_argument('--work-dir', required=True)
    parser.add_argument('--render-input', required=True, help='bt_render_input.json for structure info')
    parser.add_argument('--target', required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    params = json.loads(pathlib.Path(args.params).read_text(encoding='utf-8'))
    render_input = json.loads(pathlib.Path(args.render_input).read_text(encoding='utf-8'))
    work_dir = pathlib.Path(args.work_dir)
    wb = openpyxl.load_workbook(args.target, data_only=False)
    pk = params['pk']

    errors = []

    # CSV file → sheet mapping (order matches sheet index)
    csv_files = [
        (f'{pk}_일별수익률.csv', 0),
        (f'{pk}_월별수익률.csv', 1),
        (f'{pk}_자산종류별비중.csv', 2),
        (f'{pk}_종목별비중.csv', 3),
    ]

    for csv_name, sheet_idx in csv_files:
        csv_path = work_dir / csv_name
        if not csv_path.exists() or sheet_idx >= len(wb.sheetnames):
            continue

        # Derive data_start from render_input: find the first row with a date-type cell
        # (date columns mark CSV data rows; header/dscore rows don't have date cells)
        ri_sheet = render_input['sheets'][sheet_idx] if sheet_idx < len(render_input['sheets']) else None
        if not ri_sheet:
            continue
        data_start = None
        for row_data in sorted(ri_sheet.get('rows', []), key=lambda r: r['row_num']):
            rn = row_data['row_num']
            if rn <= 1:
                continue
            for cell_data in row_data.get('cells', []):
                if cell_data.get('value_type') == 'date':
                    data_start = rn
                    break
            if data_start:
                break
        if data_start is None:
            data_start = 2  # fallback

        df = pd.read_csv(csv_path)
        ws = wb[wb.sheetnames[sheet_idx]]
        sheet_label = f"Sheet{sheet_idx + 1}({wb.sheetnames[sheet_idx]})"
        verify_csv_sheet(ws, df, 1, data_start, errors, sheet_label)

    # Sheet 5: verify params values are placed directly (no re-derivation)
    if len(wb.sheetnames) >= 5:
        ws5 = wb[wb.sheetnames[4]]
        ri = render_input['sheets'][4] if len(render_input['sheets']) > 4 else None
        if ri:
            # Check that meta row 2 values match params
            for row_data in ri.get('rows', []):
                if row_data['row_num'] == 2:
                    for cell_data in row_data['cells']:
                        if not cell_data.get('is_formula'):
                            c = cell_data['col']
                            expected = cell_data['value']
                            actual = ws5.cell(2, c).value
                            if actual != expected:
                                try:
                                    if abs(float(actual or 0) - float(expected)) > TOLERANCE:
                                        errors.append(f"Sheet5 meta row2 col{c}: expected {expected}, got {actual}")
                                except (ValueError, TypeError):
                                    errors.append(f"Sheet5 meta row2 col{c}: expected {expected}, got {actual}")

    result = {"status": "PASS" if not errors else "FAIL", "error_count": len(errors), "errors": errors[:50]}
    pathlib.Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"Data verification: {result['status']} ({len(errors)} errors)")


if __name__ == '__main__':
    main()