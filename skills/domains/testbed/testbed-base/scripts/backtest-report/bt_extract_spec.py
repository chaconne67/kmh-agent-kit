"""Reference xlsx structure extractor. Reads any xlsx and outputs bt_spec.json."""
import argparse, json, pathlib, sys
from datetime import datetime

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')


class SafeEncoder(json.JSONEncoder):
    """Handle openpyxl descriptor types (Integer, Float, etc.) that aren't plain Python types."""
    def default(self, o):
        if isinstance(o, (int, float, bool)):
            return o
        try:
            return int(o)
        except (TypeError, ValueError):
            pass
        try:
            return float(o)
        except (TypeError, ValueError):
            pass
        return str(o)


# ---------------------------------------------------------------------------
# Style serialization helpers
# ---------------------------------------------------------------------------

def _safe_attr(obj, attr):
    """Safely get an attribute from an openpyxl descriptor, returning None on error."""
    try:
        val = getattr(obj, attr, None)
        if val is not None and isinstance(val, str) and 'Values must be' in val:
            return None
        return val
    except (TypeError, ValueError, AttributeError):
        return None


def color_to_json(color):
    """Convert openpyxl Color to JSON-serializable dict."""
    if color is None:
        return None
    theme = _safe_attr(color, 'theme')
    if theme is not None and isinstance(theme, int):
        tint = _safe_attr(color, 'tint')
        return {"theme": theme, "tint": tint if tint else 0}
    rgb = _safe_attr(color, 'rgb')
    if rgb:
        rgb = str(rgb)
        if rgb in ('00000000', '0'):
            return None
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    indexed = _safe_attr(color, 'indexed')
    if indexed is not None and isinstance(indexed, int):
        return {"indexed": indexed}
    return None


def font_to_json(font):
    if font is None:
        return None
    return {
        "name": font.name,
        "size": font.size,
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline,
        "strikethrough": font.strikethrough,
        "color": color_to_json(font.color),
    }


def fill_to_json(fill):
    if fill is None:
        return None
    return {
        "type": fill.fill_type,
        "fgColor": color_to_json(fill.fgColor) if fill.fgColor else None,
        "bgColor": color_to_json(fill.bgColor) if fill.bgColor else None,
    }


def border_side_to_json(side):
    if side is None or side.style is None:
        return None
    return {"style": side.style, "color": color_to_json(side.color)}


def border_to_json(border):
    if border is None:
        return None
    return {
        "left": border_side_to_json(border.left),
        "right": border_side_to_json(border.right),
        "top": border_side_to_json(border.top),
        "bottom": border_side_to_json(border.bottom),
    }


def alignment_to_json(alignment):
    if alignment is None:
        return None
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": alignment.wrapText,
        "text_rotation": alignment.textRotation,
    }


def cell_style_to_json(cell):
    return {
        "font": font_to_json(cell.font),
        "fill": fill_to_json(cell.fill),
        "border": border_to_json(cell.border),
        "alignment": alignment_to_json(cell.alignment),
        "number_format": cell.number_format,
    }


# ---------------------------------------------------------------------------
# Cell value extraction
# ---------------------------------------------------------------------------

def cell_value_to_json(cell):
    val = cell.value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)


def extract_cell(cell):
    if isinstance(cell, MergedCell):
        return None

    is_formula = isinstance(cell.value, str) and cell.value.startswith('=')
    return {
        "col": cell.column,
        "col_letter": get_column_letter(cell.column),
        "value": cell_value_to_json(cell),
        "is_formula": is_formula,
        "style": cell_style_to_json(cell),
    }


# ---------------------------------------------------------------------------
# Column widths / row heights
# ---------------------------------------------------------------------------

def extract_column_widths(ws):
    widths = {}
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None and dim.width > 0:
            widths[col_letter] = dim.width
    return widths


def extract_row_heights(ws):
    heights = {}
    for row_num, dim in ws.row_dimensions.items():
        if dim.height is not None:
            heights[str(row_num)] = dim.height
    return heights


# ---------------------------------------------------------------------------
# Region detection
# ---------------------------------------------------------------------------

def detect_regions(ws):
    if ws.max_row is None or ws.max_row == 0:
        return []

    non_empty_rows = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                non_empty_rows.add(cell.row)
                break

    if not non_empty_rows:
        return []

    sorted_rows = sorted(non_empty_rows)
    blocks = []
    current_block = [sorted_rows[0]]
    for r in sorted_rows[1:]:
        if r - current_block[-1] <= 2:
            current_block.append(r)
        else:
            blocks.append(current_block)
            current_block = [r]
    blocks.append(current_block)

    regions = []
    for block in blocks:
        start_row = min(block)
        end_row = max(block)
        region = extract_region(ws, start_row, end_row)
        if region:
            regions.append(region)

    return regions


def extract_region(ws, start_row, end_row):
    min_col, max_col = ws.max_column, 1
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                min_col = min(min_col, c)
                max_col = max(max_col, c)

    if max_col < min_col:
        return None

    header_end = start_row - 1
    for r in range(start_row, min(start_row + 5, end_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(min_col, max_col + 1) if ws.cell(r, c).value is not None]
        if not row_vals:
            continue
        text_count = sum(1 for v in row_vals if isinstance(v, str) and not str(v).startswith('='))
        if text_count > 0:
            header_end = r
        else:
            break

    n_header_rows = header_end - start_row + 1
    data_start = header_end + 1

    header_rows = []
    for r in range(start_row, header_end + 1):
        cells = []
        for c in range(min_col, max_col + 1):
            cell_data = extract_cell(ws.cell(r, c))
            if cell_data:
                cells.append(cell_data)
        header_rows.append(cells)

    data_sample = []
    if data_start <= end_row:
        for c in range(min_col, max_col + 1):
            cell_data = extract_cell(ws.cell(data_start, c))
            if cell_data:
                data_sample.append(cell_data)

    formulas = []
    if data_start <= end_row:
        for c in range(min_col, max_col + 1):
            val = ws.cell(data_start, c).value
            if isinstance(val, str) and val.startswith('='):
                formulas.append({
                    "col": c,
                    "col_letter": get_column_letter(c),
                    "pattern": val,
                    "style": cell_style_to_json(ws.cell(data_start, c)),
                })

    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_col": min_col,
        "end_col": max_col,
        "n_header_rows": n_header_rows,
        "data_start_row": data_start,
        "n_data_rows": max(0, end_row - data_start + 1),
        "header_rows": header_rows,
        "data_sample": data_sample,
        "formulas": formulas,
    }


# ---------------------------------------------------------------------------
# Chart extraction
# ---------------------------------------------------------------------------

def chart_color_to_json(color_obj):
    if color_obj is None:
        return None
    if hasattr(color_obj, 'schemeClr') and color_obj.schemeClr is not None:
        sc = color_obj.schemeClr
        result = {"scheme": sc.val}
        if sc.lumMod is not None:
            result["lumMod"] = sc.lumMod
        if sc.lumOff is not None:
            result["lumOff"] = sc.lumOff
        return result
    if hasattr(color_obj, 'srgbClr') and color_obj.srgbClr is not None:
        return f"#{color_obj.srgbClr.val}"
    return None


def _parse_ref(ref_obj):
    """Extract column/row info from a series reference."""
    if ref_obj is None:
        return None
    num_ref = getattr(ref_obj, 'numRef', None)
    str_ref = getattr(ref_obj, 'strRef', None)
    ref = num_ref or str_ref
    if ref is None or ref.f is None:
        return None
    # Parse "SheetName!$C$2:$C$366" → extract col letter and row range
    import re
    m = re.search(r'\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)', ref.f)
    if m:
        return {
            "min_col_letter": m.group(1),
            "min_row": int(m.group(2)),
            "max_col_letter": m.group(3),
            "max_row": int(m.group(4)),
        }
    return None


def extract_series(series):
    result = {}
    if series.graphicalProperties and series.graphicalProperties.solidFill:
        result["fill_color"] = chart_color_to_json(series.graphicalProperties.solidFill)
    if series.graphicalProperties and series.graphicalProperties.line:
        ln = series.graphicalProperties.line
        if ln.solidFill:
            result["line_color"] = chart_color_to_json(ln.solidFill)
        if ln.width is not None:
            result["line_width"] = ln.width
    # Data references
    val_ref = _parse_ref(series.val)
    cat_ref = _parse_ref(series.cat)
    if val_ref:
        result["val_ref"] = val_ref
    if cat_ref:
        result["cat_ref"] = cat_ref
    return result


def _extract_line_props(ln):
    """Extract all LineProperties to JSON."""
    if ln is None:
        return None
    result = {"width": ln.width, "noFill": ln.noFill, "prstDash": ln.prstDash}
    if ln.solidFill:
        result["solidFill"] = chart_color_to_json(ln.solidFill)
    return result


def _extract_graphical_props(gp):
    """Extract GraphicalProperties (border + fill) to JSON."""
    if gp is None:
        return None
    result = {}
    if gp.ln:
        result["line"] = _extract_line_props(gp.ln)
    if gp.solidFill:
        result["solidFill"] = chart_color_to_json(gp.solidFill)
    if gp.noFill:
        result["noFill"] = True
    return result if result else None


def _extract_gridlines(gl):
    """Extract ChartLines (gridlines) to JSON."""
    if gl is None:
        return None
    if gl.spPr and gl.spPr.ln:
        return _extract_line_props(gl.spPr.ln)
    return {}


def _serialize_xml(obj):
    """Serialize any openpyxl Serialisable object to XML string."""
    if obj is None:
        return None
    try:
        from lxml import etree
        return etree.tostring(obj.to_tree(), encoding='unicode')
    except Exception:
        return None


def _extract_axis(axis):
    """Extract axis as raw XML — captures 100% of properties."""
    if axis is None:
        return {}
    return {
        "_xml": _serialize_xml(axis),
        "_type": type(axis).__name__,
    }


def extract_chart(chart):
    """Extract chart as full XML + anchor/series metadata for data remapping.
    XML captures 100% of formatting. Series refs are extracted separately
    so the assembler can remap data ranges for the target file."""
    from lxml import etree

    chart_type = type(chart).__name__

    # Full chart XML — captures ALL properties
    chart_xml = None
    try:
        tree = chart._write()
        chart_xml = etree.tostring(tree, encoding='unicode')
    except Exception:
        pass

    # Anchor (not in chart XML — stored in the drawing)
    anchor = None
    anchor_detail = None
    if hasattr(chart, 'anchor') and chart.anchor is not None:
        try:
            marker = chart.anchor._from
            anchor = f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
            anchor_detail = {
                "col": marker.col, "row": marker.row,
                "colOff": marker.colOff or 0,
                "rowOff": marker.rowOff or 0,
            }
        except (AttributeError, TypeError):
            pass

    return {
        "type": chart_type,
        "width": chart.width,
        "height": chart.height,
        "ref_anchor": anchor,
        "ref_anchor_detail": anchor_detail,
        # Full chart XML — single source of truth for all formatting
        "_chart_xml": chart_xml,
        # Series refs extracted for data remapping by assembler
        "series": [extract_series(s) for s in chart.series],
    }


# ---------------------------------------------------------------------------
# Sheet / workbook extraction
# ---------------------------------------------------------------------------

def extract_sheet(ws, index):
    return {
        "index": index,
        "name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "column_widths": extract_column_widths(ws),
        "row_heights": extract_row_heights(ws),
        "merged_cells": [str(m) for m in ws.merged_cells.ranges],
        "regions": detect_regions(ws),
        "charts": [extract_chart(c) for c in ws._charts],
    }


def extract_workbook(wb):
    return {"sheets": [extract_sheet(wb[name], idx) for idx, name in enumerate(wb.sheetnames)]}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Extract reference xlsx structure to JSON')
    parser.add_argument('--ref', required=True, help='Path to reference xlsx')
    parser.add_argument('--output', required=True, help='Output bt_spec.json path')
    args = parser.parse_args()

    ref_path = pathlib.Path(args.ref)
    wb = openpyxl.load_workbook(ref_path, data_only=False)
    spec = extract_workbook(wb)

    out_path = pathlib.Path(args.output)
    out_path.write_text(json.dumps(spec, ensure_ascii=False, indent=2, cls=SafeEncoder), encoding='utf-8')
    print(f'Spec extracted: {len(spec["sheets"])} sheets -> {out_path}')


if __name__ == '__main__':
    main()