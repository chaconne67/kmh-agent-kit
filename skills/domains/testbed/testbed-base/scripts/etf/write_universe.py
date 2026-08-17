# -*- coding: utf-8 -*-
"""
universe.json → 엑셀 기입 (순수 렌더러)

사용법:
  python write_universe.py universe.json --output-xlsx output.xlsx

- JSON의 metadata에서 로컬 원본 경로를 읽음
- JSON의 items[].final 값을 엑셀에 기입
- 원본은 바꾸지 않고 새 xlsx에 저장
- 하드코딩 없음: 경로, 시트, 열 매핑 모두 JSON에서 결정
"""

import argparse
import io
import json
import sys
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl


def main():
    parser = argparse.ArgumentParser(description="universe.json → 엑셀 기입")
    parser.add_argument("json_path", help="universe.json 경로")
    parser.add_argument("--output-xlsx", required=True, help="새 xlsx 출력 경로")
    parser.add_argument("--dry-run", action="store_true", help="변경사항만 출력, 실제 저장 안함")
    args = parser.parse_args()

    json_path = Path(args.json_path).resolve()
    with json_path.open("r", encoding="utf-8") as f:
        universe = json.load(f)

    meta = universe["metadata"]
    source_path = Path(meta["source_file"])
    if not source_path.is_absolute():
        source_path = json_path.parent / source_path
    source_path = source_path.resolve()
    output_path = Path(args.output_xlsx).resolve()

    if not source_path.is_file():
        print(f"ERROR: 로컬 원본 xlsx가 없습니다: {source_path}", file=sys.stderr)
        sys.exit(1)
    if source_path == output_path:
        print("ERROR: 원본과 출력 경로가 같습니다.", file=sys.stderr)
        sys.exit(1)
    if output_path.suffix.lower() != ".xlsx":
        print("ERROR: 출력 파일은 .xlsx여야 합니다.", file=sys.stderr)
        sys.exit(1)
    if not output_path.parent.is_dir():
        print(f"ERROR: 출력 디렉토리가 없습니다: {output_path.parent}", file=sys.stderr)
        sys.exit(1)
    if output_path.exists() and not args.dry_run:
        print(f"ERROR: 출력 파일이 이미 있습니다: {output_path}", file=sys.stderr)
        sys.exit(1)

    print(f"[원본] {source_path}", file=sys.stderr)
    print(f"[출력] {output_path}", file=sys.stderr)

    wb = openpyxl.load_workbook(source_path)
    ws = wb.worksheets[0]
    print(f"[시트] {ws.title}", file=sys.stderr)

    # 열 매핑: metadata.col_mapping에서 읽음 (필수)
    if not meta.get("col_mapping"):
        print("ERROR: metadata.col_mapping이 없거나 null입니다. §2에서 col_mapping을 먼저 채우세요.", file=sys.stderr)
        sys.exit(1)
    col_map = {k: int(v) for k, v in meta["col_mapping"].items()}

    # risk_grade 타입 검증: 숫자이면 경고 후 중단
    for item in universe["items"]:
        rg = (item.get("final") or {}).get("risk_grade")
        if rg is not None and isinstance(rg, (int, float)):
            print(f"ERROR: 행{item['row']} risk_grade가 숫자({rg})입니다. "
                  f"텍스트 레이블(예: '높은위험')이어야 합니다.", file=sys.stderr)
            sys.exit(1)

    changes = []
    for item in universe["items"]:
        row = item["row"]
        final = item.get("final", {})
        if not final:
            continue

        for field, col in col_map.items():
            new_val = final.get(field)
            if new_val is None:
                continue
            old_val = ws.cell(row=row, column=col).value
            if str(old_val) != str(new_val):
                changes.append((row, field, old_val, new_val, col))
                if not args.dry_run:
                    ws.cell(row=row, column=col, value=new_val)

    # 결과 보고
    print(f"\n[변경] {len(changes)}건", file=sys.stderr)
    for row, field, old, new, _ in changes:
        print(f"  행{row} {field}: {old} → {new}", file=sys.stderr)

    if args.dry_run:
        print("\n[dry-run] 저장하지 않음", file=sys.stderr)
    else:
        wb.save(output_path)
        print(f"\n[저장] {output_path}", file=sys.stderr)

    # JSON 결과 출력 (stdout)
    result = {
        "status": "dry_run" if args.dry_run else "saved",
        "source": str(source_path),
        "file": str(output_path),
        "total_items": len(universe["items"]),
        "changes": len(changes),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
