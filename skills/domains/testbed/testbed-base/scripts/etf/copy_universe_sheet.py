# -*- coding: utf-8 -*-
"""§8: source_file → target_file의 지정 시트에 덮어쓰기 (순수 렌더러)

사용법:
  python copy_universe_sheet.py universe.json --output-xlsx output.xlsx

- JSON metadata에서 로컬 source_file, target_file, target_sheet를 읽음
- 소스 파일의 첫 번째 시트 전체를 타겟 파일의 지정 시트에 덮어쓰기
- 타겟 원본은 바꾸지 않고 새 xlsx에 저장
- 하드코딩 없음: 경로, 파일명, 시트명 모두 JSON에서 결정
"""
import argparse, json, io, sys
from copy import copy
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl


def main():
    parser = argparse.ArgumentParser(description="source_file → target_file 시트 덮어쓰기")
    parser.add_argument("json_path", help="universe.json 경로")
    parser.add_argument("--output-xlsx", required=True, help="새 xlsx 출력 경로")
    parser.add_argument("--dry-run", action="store_true", help="변경사항만 출력, 실제 저장 안함")
    args = parser.parse_args()

    json_path = Path(args.json_path).resolve()
    with json_path.open("r", encoding="utf-8") as f:
        u = json.load(f)

    meta = u["metadata"]
    src_path = Path(meta["source_file"])
    tgt_path = Path(meta["target_file"])
    if not src_path.is_absolute():
        src_path = json_path.parent / src_path
    if not tgt_path.is_absolute():
        tgt_path = json_path.parent / tgt_path
    src_path = src_path.resolve()
    tgt_path = tgt_path.resolve()
    output_path = Path(args.output_xlsx).resolve()

    for label, path in (("소스", src_path), ("타겟", tgt_path)):
        if not path.is_file():
            print(f"ERROR: 로컬 {label} xlsx가 없습니다: {path}", file=sys.stderr)
            sys.exit(1)
    if output_path in {src_path, tgt_path}:
        print("ERROR: 출력 경로가 원본 경로와 같습니다.", file=sys.stderr)
        sys.exit(1)
    if output_path.suffix.lower() != ".xlsx":
        print("ERROR: 출력 파일은 .xlsx여야 합니다.", file=sys.stderr)
        sys.exit(1)
    if not output_path.parent.is_dir():
        print(f"ERROR: 출력 디렉토리가 없습니다: {output_path.parent}", file=sys.stderr)
        sys.exit(1)
    if output_path.exists() and not args.dry_run:
        print(f"ERROR: 출력 파일이 이미 있습니다: {output_path}", file=sys.stderr)
        sys.exit(1)

    tgt_sheet = meta.get("target_sheet")
    if not tgt_sheet:
        print("ERROR: metadata.target_sheet가 없거나 null입니다. §1에서 target_sheet를 먼저 결정하세요.", file=sys.stderr)
        sys.exit(1)

    print(f"[소스] {src_path}", file=sys.stderr)
    print(f"[타겟 원본] {tgt_path} → 시트: {tgt_sheet}", file=sys.stderr)
    print(f"[출력] {output_path}", file=sys.stderr)

    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb.worksheets[0]
    print(f"[소스 시트] {src_ws.title} ({src_ws.max_row}행 x {src_ws.max_column}열)", file=sys.stderr)

    tgt_wb = openpyxl.load_workbook(tgt_path)

    # JSON에서 지정한 시트명으로 찾기
    if tgt_sheet not in tgt_wb.sheetnames:
        print(f"ERROR: '{tgt_sheet}' 시트를 찾을 수 없습니다.", file=sys.stderr)
        for name in tgt_wb.sheetnames:
            print(f"  - {name}", file=sys.stderr)
        sys.exit(1)

    tgt_ws = tgt_wb[tgt_sheet]
    print(f"[타겟 시트] {tgt_ws.title} (기존 {tgt_ws.max_row}행 x {tgt_ws.max_column}열)", file=sys.stderr)

    # 기존 데이터+서식 클리어 (source보다 target이 클 수 있으므로 max 범위로 clear)
    clear_rows = max(src_ws.max_row, tgt_ws.max_row)
    clear_cols = max(src_ws.max_column, tgt_ws.max_column)
    from openpyxl.styles import Font, PatternFill, Border, Alignment
    default_font = Font()
    default_fill = PatternFill()
    default_border = Border()
    default_alignment = Alignment()
    for row in tgt_ws.iter_rows(min_row=1, max_row=clear_rows, max_col=clear_cols):
        for cell in row:
            cell.value = None
            # 소스 범위 밖의 셀은 서식도 초기화
            if cell.row > src_ws.max_row or cell.column > src_ws.max_column:
                cell.font = copy(default_font)
                cell.fill = copy(default_fill)
                cell.border = copy(default_border)
                cell.alignment = copy(default_alignment)
                cell.number_format = "General"

    # 소스 → 타겟 복사 (값 + 서식)
    for row_idx, row in enumerate(src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column), 1):
        for col_idx, src_cell in enumerate(row, 1):
            tgt_cell = tgt_ws.cell(row=row_idx, column=col_idx, value=src_cell.value)
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.number_format = src_cell.number_format

    # 열 너비 복사
    for col_letter, dim in src_ws.column_dimensions.items():
        tgt_ws.column_dimensions[col_letter].width = dim.width

    # 행 높이 복사
    for row_idx, dim in src_ws.row_dimensions.items():
        tgt_ws.row_dimensions[row_idx].height = dim.height

    if args.dry_run:
        print(f"\n[dry-run] {src_ws.max_row}행 x {src_ws.max_column}열 복사 예정, 저장하지 않음", file=sys.stderr)
    else:
        tgt_wb.save(output_path)
        print(f"\n[완료] {src_ws.max_row}행 x {src_ws.max_column}열 → {tgt_ws.title} 덮어쓰기 저장", file=sys.stderr)

    result = {
        "status": "dry_run" if args.dry_run else "saved",
        "source": str(src_path),
        "target": str(tgt_path),
        "output": str(output_path),
        "target_sheet": tgt_sheet,
        "rows": src_ws.max_row,
        "cols": src_ws.max_column,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
